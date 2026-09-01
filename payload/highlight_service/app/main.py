from __future__ import annotations

import asyncio
import csv
import io
import json
import os
import re
import shutil
import subprocess
import zipfile
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Literal
from urllib.parse import quote, urlparse
from uuid import uuid4

import httpx
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from starlette.background import BackgroundTask

from .config import settings
from .db import Database, json_field, utc_now
from .guardian import Guardian
from .media import MediaError, MediaTools, evenly_timed_captions
from .pipeline import HighlightPipeline, safe_id
from .room_backup import export_room_backup, import_room_backup
from .rooms import RoomRegistry
from .text_normalize import simplify_value, to_simplified


db = Database(settings.db_path)
pipeline = HighlightPipeline(settings, db)
media = pipeline.media
guardian = Guardian(settings, db, pipeline)
templates = Jinja2Templates(directory=str(settings.service_root / "app" / "templates"))


def current_app_version() -> str:
    version_file = settings.service_root.parent / "VERSION"
    try:
        return version_file.read_text(encoding="utf-8").strip() or "未标记"
    except FileNotFoundError:
        return "开发版"


def _version_parts(value: str) -> tuple[int, ...]:
    try:
        return tuple(int(part) for part in value.strip().lstrip("vV").split("."))
    except ValueError:
        return (0,)


def github_update_status() -> dict[str, Any]:
    root = settings.service_root.parent
    config_path = root / "update-config.json"
    if not config_path.is_file():
        return {"ok": False, "current_version": current_app_version(), "message": "缺少更新配置文件"}
    try:
        config = json.loads(config_path.read_text(encoding="utf-8-sig"))
        repository = str(config.get("repository") or "").strip()
        if repository.count("/") != 1 or any(char not in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_./" for char in repository):
            raise ValueError("更新仓库地址无效")
        release_url = f"https://github.com/{repository}/releases/latest"
        try:
            response = httpx.get(
                f"https://api.github.com/repos/{repository}/releases/latest",
                headers={"Accept": "application/vnd.github+json", "User-Agent": "Live-Highlight-Web-Updater"},
                timeout=20,
                follow_redirects=True,
            )
            response.raise_for_status()
            release = response.json()
            available = str(release.get("tag_name") or "").strip().lstrip("vV")
            release_url = str(release.get("html_url") or release_url)
        except httpx.HTTPError:
            response = httpx.get(
                f"https://raw.githubusercontent.com/{repository}/main/payload/VERSION",
                headers={"User-Agent": "Live-Highlight-Web-Updater"}, timeout=20, follow_redirects=True,
            )
            response.raise_for_status()
            available = response.text.strip().lstrip("vV")
        current = current_app_version()
        return {
            "ok": True,
            "current_version": current,
            "available_version": available,
            "update_available": _version_parts(available) > _version_parts(current),
            "release_url": release_url,
            "message": "发现新版本" if _version_parts(available) > _version_parts(current) else "当前已经是最新版本",
        }
    except (OSError, ValueError, json.JSONDecodeError, httpx.HTTPError) as exc:
        return {"ok": False, "current_version": current_app_version(), "message": f"检查更新失败：{exc}"}


@asynccontextmanager
async def lifespan(_: FastAPI):
    pid_file = settings.data_dir / "service.pid"
    pid_file.write_text(str(os.getpid()), encoding="ascii")
    pipeline.start()
    async def retention_loop() -> None:
        while True:
            cleanup_expired_candidate_media()
            cleanup_all_ready_segments()
            await asyncio.sleep(3600)
    retention_task = asyncio.create_task(retention_loop())
    try:
        yield
    finally:
        retention_task.cancel()
        try:
            await retention_task
        except asyncio.CancelledError:
            pass
        pipeline.stop()
        try:
            if pid_file.read_text(encoding="ascii").strip() == str(os.getpid()):
                pid_file.unlink()
        except FileNotFoundError:
            pass


app = FastAPI(title="直播高光审核台", version="0.1.0", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=str(settings.service_root / "app" / "static")), name="static")
app.title = "直播高光审核台"


def validate_ranges(raw_ranges: list[dict[str, Any]]) -> tuple[list[dict[str, float]], float]:
    if not 1 <= len(raw_ranges) <= settings.max_source_ranges:
        raise HTTPException(422, f"保留区间必须为 1–{settings.max_source_ranges} 段")
    ranges = [{"start": float(item["start"]), "end": float(item["end"])} for item in raw_ranges]
    ranges.sort(key=lambda item: item["start"])
    # Model-selected clauses can legitimately leave a short bridge fragment.
    # FFmpeg can render it, so only reject empty or near-zero ranges here.
    if any(item["end"] <= item["start"] or item["end"] - item["start"] < 0.05 for item in ranges):
        raise HTTPException(422, "保留区间必须有效且不能短于 0.05 秒")
    if any(ranges[index]["start"] < ranges[index - 1]["end"] for index in range(1, len(ranges))):
        raise HTTPException(422, "保留区间不能重叠")
    if ranges[-1]["end"] - ranges[0]["start"] > settings.max_source_span_seconds:
        raise HTTPException(422, f"所有保留句必须来自同一段 {settings.max_source_span_seconds:.0f} 秒内容")
    duration = sum(item["end"] - item["start"] for item in ranges)
    if not settings.clip_min_seconds <= duration <= settings.clip_max_seconds:
        raise HTTPException(422, f"成片必须为 {settings.clip_min_seconds:.0f}–{settings.clip_max_seconds:.0f} 秒")
    return ranges, duration


def candidate_view(row: dict[str, Any]) -> dict[str, Any]:
    result = simplify_value(dict(row))
    result["captions"] = json_field(result, "captions_json", [])
    result["risks"] = json_field(result, "risk_json", [])
    result["keyframes"] = json_field(result, "keyframes_json", [])
    result["source_ranges"] = json_field(result, "source_ranges_json", [])
    result["kept_clauses"] = json_field(result, "kept_clauses_json", [])
    result["removed_clauses"] = json_field(result, "removed_clauses_json", [])
    result["compliance_hits"] = json_field(result, "compliance_hits_json", [])
    result["duration"] = round(sum(float(r["end"]) - float(r["start"]) for r in result["source_ranges"]), 2)
    if not result["source_ranges"]:
        result["duration"] = round(float(result["end_time"]) - float(result["start_time"]), 2)
    result["source_ranges_text"] = json.dumps(result["source_ranges"], ensure_ascii=False)
    result["caption_text"] = "\n".join(item.get("text", "") for item in result["captions"])
    analysis_version = str(result.get("analysis_version") or "").lower()
    if "deepseek" in analysis_version:
        result["model_role"] = "DeepSeek 补漏"
    elif "gpt-5.4-mini" in analysis_version:
        result["model_role"] = "GPT-5.4-mini 备用主选"
    elif "gpt-5.4" in analysis_version:
        result["model_role"] = "GPT-5.4 主选"
    else:
        result["model_role"] = "GPT-5.5 主选"
    try:
        created = datetime.fromisoformat(str(result.get("created_at") or "").replace("Z", "+00:00"))
        if created.tzinfo is None:
            created = created.replace(tzinfo=timezone.utc)
        result["created_at_local"] = created.astimezone(timezone(timedelta(hours=8))).strftime("%Y-%m-%d %H:%M")
    except ValueError:
        result["created_at_local"] = str(result.get("created_at") or "")
    result["exported_at_local"] = ""
    if result.get("exported_at"):
        try:
            exported = datetime.fromisoformat(str(result["exported_at"]).replace("Z", "+00:00"))
            if exported.tzinfo is None:
                exported = exported.replace(tzinfo=timezone.utc)
            result["exported_at_local"] = exported.astimezone(
                timezone(timedelta(hours=8))
            ).strftime("%Y-%m-%d %H:%M")
        except ValueError:
            result["exported_at_local"] = str(result["exported_at"])
    return result


def room_live_state(room: dict[str, Any]) -> tuple[str, str]:
    status = str(room.get("live_status") or "unknown")
    if status == "live":
        return "live", "直播中"
    if status == "offline":
        return "offline", "未开播"
    if room.get("live_checked_at"):
        return "unknown", "状态未知"
    return "checking", "检测中"


def _clock_text(seconds: float) -> str:
    value = max(0, int(seconds))
    hours, remainder = divmod(value, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}" if hours else f"{minutes:02d}:{secs:02d}"


def candidate_output_date_range(value: str) -> tuple[str, str] | None:
    value = value.strip()
    if not value:
        return None
    try:
        china_tz = timezone(timedelta(hours=8))
        local_start = datetime.strptime(value, "%Y-%m-%d").replace(tzinfo=china_tz)
    except ValueError as exc:
        raise HTTPException(422, "候选产出日期格式不正确") from exc
    utc_start = local_start.astimezone(timezone.utc)
    utc_end = (local_start + timedelta(days=1)).astimezone(timezone.utc)
    return utc_start.isoformat(timespec="seconds"), utc_end.isoformat(timespec="seconds")


def optional_query_int(value: str | int | None, label: str) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise HTTPException(422, f"{label}筛选值无效") from exc


def candidate_local_date(candidate: dict[str, Any], compact: bool = False) -> str:
    """Return a candidate's output day in Beijing time for user-facing names."""
    try:
        created = datetime.fromisoformat(str(candidate.get("created_at") or "").replace("Z", "+00:00"))
        if created.tzinfo is None:
            created = created.replace(tzinfo=timezone.utc)
        pattern = "%Y%m%d" if compact else "%Y-%m-%d"
        return created.astimezone(timezone(timedelta(hours=8))).strftime(pattern)
    except ValueError:
        return datetime.now(timezone(timedelta(hours=8))).strftime("%Y%m%d" if compact else "%Y-%m-%d")


def candidate_room_file_prefix(candidate: dict[str, Any]) -> str:
    raw_sequence = str(candidate.get("room_sequence") or "").strip()
    sequence = safe_id(raw_sequence) if raw_sequence else ""
    room_name = safe_id(str(candidate.get("room_name") or candidate.get("source_id") or "未知直播间"))
    return f"{sequence}_{room_name}" if sequence else room_name


def candidate_download_filename(candidate: dict[str, Any]) -> str:
    return f"{candidate_room_file_prefix(candidate)}_{candidate_local_date(candidate, True)}_素材{int(candidate['id']):04d}.mp4"


def active_recording_state(room: dict[str, Any]) -> dict[str, Any]:
    result = {
        "recording_active": False,
        "active_file": "",
        "active_started_at": "",
        "active_elapsed_text": "",
        "active_size_mb": 0.0,
        "active_progress": 0.0,
        "active_remaining_text": "",
        "write_fresh": False,
    }
    try:
        room_key = str(room.get("source_key") or safe_id(room["name"]))
        latest = max(
            (path for path in settings.input_dir.rglob("*")
             if path.is_file()
             and path.suffix.lower() in {".ts", ".mp4", ".mkv", ".flv"}
             and safe_id(path.parent.name) == room_key),
            key=lambda path: path.stat().st_mtime,
            default=None,
        )
        if latest is None:
            return result
        stat = latest.stat()
        now_ts = datetime.now().timestamp()
        elapsed = max(0.0, now_ts - stat.st_ctime)
        fresh = now_ts - stat.st_mtime < 45
        target = float(settings.recorder_segment_seconds)
        result.update({
            "recording_active": bool(fresh and room.get("enabled")),
            "active_file": latest.name,
            "active_started_at": datetime.fromtimestamp(stat.st_ctime).strftime("%H:%M:%S"),
            "active_elapsed_text": _clock_text(elapsed),
            "active_size_mb": round(stat.st_size / 1024 / 1024, 1),
            "active_progress": round(min(elapsed / target * 100, 100), 1),
            "active_remaining_text": _clock_text(max(0, target - elapsed)),
            "write_fresh": fresh,
        })
    except OSError:
        pass
    return result


def room_processing_state(room: dict[str, Any]) -> tuple[str, str]:
    counts = {row["status"]: int(row["count"]) for row in db.all(
        "SELECT status,COUNT(*) AS count FROM recording_segments WHERE room_id=? AND status<>'cleaned' GROUP BY status", (room["id"],)
    )}
    for status, key, label in (
        ("awaiting_finalization", "queued", "等待录像封口"),
        ("transcribing", "transcribing", "转写中"),
        ("gpt_analyzing", "gpt", "GPT分析中"),
        ("deepseek_analyzing", "deepseek", "DeepSeek补漏中"),
        ("transcribed", "queued", "等待模型分析"),
        ("ai_waiting", "retry", "模型重试中"),
        ("ai_retry_paused", "error", "模型重试已暂停"),
        ("ai_abandoned", "paused", "模型任务已舍弃"),
        ("discovered", "queued", "等待处理"),
        ("rendering", "rendering", "渲染中"),
    ):
        if counts.get(status):
            return key, label
    if counts.get("error") or counts.get("asr_unavailable") or room.get("last_error"):
        return "error", "处理异常"
    if sum(counts.values()):
        return "complete", "处理完成"
    return "idle", "暂无素材"


def room_cards() -> list[dict[str, Any]]:
    rows = db.all(
        """SELECT r.*,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.room_id=r.id
              AND s.status NOT IN ('awaiting_finalization','cleaned')) AS segment_total,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.room_id=r.id AND s.transcribed_at<>'' AND s.status<>'cleaned') AS segment_transcribed,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.room_id=r.id AND s.model_submitted_at<>'' AND s.status<>'cleaned') AS segment_model_submitted,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.room_id=r.id AND s.analyzed_at<>'' AND s.status<>'cleaned') AS segment_analyzed,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.room_id=r.id AND s.status='discovered') AS waiting_asr,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.room_id=r.id AND s.status='transcribing') AS active_asr,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.room_id=r.id AND s.status IN ('transcribed','ai_waiting','ai_retry_paused')) AS waiting_ai,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.room_id=r.id AND s.status IN ('gpt_analyzing','deepseek_analyzing')) AS active_ai,
           (SELECT COUNT(*) FROM highlight_candidates h WHERE h.room_id=r.id AND h.status='visual_review'
              AND EXISTS (SELECT 1 FROM recording_segments s WHERE s.session_id=h.session_id
                AND NOT(s.timeline_end<=h.start_time OR s.timeline_start>=h.end_time))
              AND NOT EXISTS (SELECT 1 FROM recording_segments s WHERE s.session_id=h.session_id
                AND NOT(s.timeline_end<=h.start_time OR s.timeline_start>=h.end_time)
                AND s.status NOT IN ('complete','analyzed'))) AS waiting_render,
           (SELECT COUNT(*) FROM highlight_candidates h WHERE h.room_id=r.id AND h.status='rendering') AS active_render,
           (SELECT COUNT(*) FROM highlight_candidates h WHERE h.room_id=r.id AND h.status='render_error') AS render_error_count,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.room_id=r.id AND s.analyzed_at<>'' AND s.status<>'cleaned') AS segment_processed,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.room_id=r.id AND s.status IN ('error','asr_unavailable','ai_retry_paused')) AS segment_errors,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.room_id=r.id AND s.status<>'cleaned' AND EXISTS(
              SELECT 1 FROM highlight_candidates h WHERE h.session_id=s.session_id
                AND h.status NOT IN ('superseded','rejected','render_error')
                AND NOT(h.end_time<=s.timeline_start OR h.start_time>=s.timeline_end))) AS segment_yielded,
           (SELECT COUNT(*) FROM highlight_candidates h WHERE h.room_id=r.id AND h.status NOT IN ('superseded','render_error') AND h.media_cleaned_at='') AS candidate_total,
           (SELECT COUNT(*) FROM highlight_candidates h WHERE h.room_id=r.id AND h.status='pending_review') AS pending_review,
           (SELECT COUNT(*) FROM highlight_candidates h WHERE h.room_id=r.id AND h.status='accepted') AS accepted_count,
           (SELECT COUNT(*) FROM highlight_candidates h WHERE h.room_id=r.id AND h.status='rejected') AS rejected_count,
           (SELECT COUNT(*) FROM highlight_candidates h WHERE h.room_id=r.id AND h.status='exported' AND h.media_cleaned_at='') AS exported_count
           FROM live_rooms r
           WHERE r.archived=0 ORDER BY r.sequence,r.id"""
    )
    for row in rows:
        row["live_state"], row["live_label"] = room_live_state(row)
        row.update(active_recording_state(row))
        if row["recording_active"]:
            row["recording_state"], row["recording_label"] = "active", "正在录制"
        elif row.get("enabled"):
            row["recording_state"], row["recording_label"] = "enabled", "录制已开"
        else:
            row["recording_state"], row["recording_label"] = "paused", "录制暂停"
        row["processing_state"], row["processing_label"] = room_processing_state(row)
        row["state"] = row["processing_label"]
        row["effective_count"] = int(row["accepted_count"]) + int(row["exported_count"])
    return rows


def get_candidate(candidate_id: int) -> dict[str, Any]:
    row = db.one(
        """SELECT c.*, i.internal_code, i.name AS catalog_name,r.sequence AS room_sequence,r.name AS room_name,
                  i.qianchuan_product_id, i.qianchuan_plan_id
           FROM highlight_candidates c LEFT JOIN catalog_items i ON i.id=c.catalog_item_id
           LEFT JOIN live_rooms r ON r.id=c.room_id
           WHERE c.id=?""",
        (candidate_id,),
    )
    if not row:
        raise HTTPException(404, "候选片段不存在")
    return candidate_view(row)


def ensure_publish_job(candidate: dict[str, Any]) -> dict[str, Any]:
    existing = db.one("SELECT * FROM publish_jobs WHERE candidate_id=?", (candidate["id"],))
    complete = bool(candidate.get("internal_code") and candidate.get("qianchuan_product_id") and candidate.get("qianchuan_plan_id"))
    status = "ready" if complete else "missing_info"
    now = utc_now()
    if existing:
        if existing["status"] not in {"exported", "published", "cancelled"}:
            db.execute(
                """UPDATE publish_jobs SET room_id=?,catalog_item_id=?,internal_code_snapshot=?,
                   qianchuan_product_id_snapshot=?,qianchuan_plan_id_snapshot=?,status=?,error='',updated_at=? WHERE id=?""",
                (candidate.get("room_id"), candidate.get("catalog_item_id"), candidate.get("internal_code") or "",
                 candidate.get("qianchuan_product_id") or "", candidate.get("qianchuan_plan_id") or "",
                 status, now, existing["id"]),
            )
        return db.one("SELECT * FROM publish_jobs WHERE id=?", (existing["id"],)) or existing
    job_id = db.execute(
        """INSERT INTO publish_jobs
           (candidate_id,room_id,catalog_item_id,internal_code_snapshot,qianchuan_product_id_snapshot,
            qianchuan_plan_id_snapshot,status,created_at,updated_at)
           VALUES(?,?,?,?,?,?,?,?,?)""",
        (candidate["id"], candidate.get("room_id"), candidate.get("catalog_item_id"),
         candidate.get("internal_code") or "", candidate.get("qianchuan_product_id") or "",
         candidate.get("qianchuan_plan_id") or "", status, now, now),
    )
    return db.one("SELECT * FROM publish_jobs WHERE id=?", (job_id,)) or {}


def allowed_media(path_text: str) -> Path:
    path = Path(path_text).resolve()
    roots = [settings.output_dir.resolve(), settings.keyframe_dir.resolve()]
    if not any(path == root or root in path.parents for root in roots):
        raise HTTPException(403, "不允许访问该文件")
    if not path.exists() or not path.is_file():
        raise HTTPException(404, "文件不存在")
    return path


def _candidate_overlaps_segment(candidate: dict[str, Any], segment: dict[str, Any]) -> bool:
    ranges = json_field(candidate, "source_ranges_json", []) or [{
        "start": candidate["start_time"], "end": candidate["end_time"],
    }]
    return any(
        float(item["end"]) > float(segment["timeline_start"])
        and float(item["start"]) < float(segment["timeline_end"])
        for item in ranges
    )


def segment_cleanup_report(segment_id: int) -> dict[str, Any]:
    segment = db.one("SELECT * FROM recording_segments WHERE id=?", (segment_id,))
    if not segment:
        raise HTTPException(404, "录像分片不存在")
    candidates = [
        candidate for candidate in db.all(
            "SELECT * FROM highlight_candidates WHERE session_id=? ORDER BY id", (segment["session_id"],)
        ) if _candidate_overlaps_segment(candidate, segment)
        and candidate["status"] != "superseded"
    ]
    blockers: list[str] = []
    dispositions: list[dict[str, Any]] = []
    for candidate in candidates:
        status = candidate["status"]
        disposition = ""
        if status == "rejected":
            disposition = "已明确不需要"
        elif status == "exported":
            output_path = Path(candidate["output_path"]) if candidate.get("output_path") else None
            if candidate.get("media_cleaned_at"):
                disposition = "已导出，成片已按保留期清理"
            elif output_path and output_path.is_file():
                disposition = "成片已导出"
            else:
                blockers.append(f"候选 #{candidate['id']} 的导出成片不存在")
                disposition = "需要重新导出"
        elif status == "deferred":
            blockers.append(f"候选 #{candidate['id']} 正在暂存待用")
            disposition = "暂存待用"
        elif status == "accepted":
            blockers.append(f"候选 #{candidate['id']} 已接受但尚未导出成片")
            disposition = "等待导出"
        elif status == "render_error":
            blockers.append(f"候选 #{candidate['id']} 渲染异常，必须保留原录像以便重试")
            disposition = "渲染异常待恢复"
        else:
            blockers.append(f"候选 #{candidate['id']} 尚未明确去向")
            disposition = "待审核"
        dispositions.append({"candidate_id": candidate["id"], "status": status, "disposition": disposition})
    if segment["status"] not in {"complete", "analyzed", "cleaned"}:
        blockers.append(f"分片仍处于“{segment['status']}”阶段")
    source_path = Path(segment["path"])
    estimated_bytes = source_path.stat().st_size if source_path.is_file() else 0
    estimated_paths: set[str] = set()
    for candidate in candidates:
        for field in ("preview_path", "output_path"):
            path_text = candidate.get(field) or ""
            path = Path(path_text) if path_text else None
            if candidate["status"] != "exported" and path and path.is_file() and str(path.resolve()) not in estimated_paths:
                estimated_bytes += path.stat().st_size
                estimated_paths.add(str(path.resolve()))
        for path in settings.keyframe_dir.glob(f"candidate_{candidate['id']}_*"):
            if path.is_file():
                estimated_bytes += path.stat().st_size
    return {
        "segment": segment,
        "candidates": dispositions,
        "candidate_count": len(candidates),
        "blockers": blockers,
        "cleanable": not blockers and segment["status"] != "cleaned",
        "already_cleaned": segment["status"] == "cleaned",
        "estimated_bytes": estimated_bytes,
        "estimated_gb": round(estimated_bytes / (1024 ** 3), 3),
    }


def _unlink_cleanup_file(path: Path, roots: list[Path]) -> int:
    try:
        resolved = path.resolve()
        if not any(resolved == root or root in resolved.parents for root in roots):
            return 0
        if not resolved.is_file():
            return 0
        size = resolved.stat().st_size
        resolved.unlink()
        return size
    except (FileNotFoundError, OSError):
        return 0


def cleanup_candidate_media(candidate_id: int) -> dict[str, Any]:
    candidate = db.one("SELECT * FROM highlight_candidates WHERE id=?", (candidate_id,))
    if not candidate:
        raise HTTPException(404, "候选素材不存在")
    if candidate.get("media_cleaned_at"):
        return {"candidate_id": candidate_id, "released_bytes": candidate.get("media_released_bytes", 0),
                "idempotent": True}
    roots = [settings.output_dir.resolve()]
    released = 0
    seen: set[str] = set()
    for field in ("preview_path", "output_path"):
        path_text = str(candidate.get(field) or "")
        if not path_text:
            continue
        resolved_text = str(Path(path_text).resolve())
        if resolved_text in seen:
            continue
        released += _unlink_cleanup_file(Path(path_text), roots)
        seen.add(resolved_text)
    now = utc_now()
    db.execute(
        """UPDATE highlight_candidates SET preview_path='',output_path='',media_cleaned_at=?,
           media_released_bytes=?,updated_at=? WHERE id=?""",
        (now, released, now, candidate_id),
    )
    return {"candidate_id": candidate_id, "released_bytes": released, "idempotent": False}


def cleanup_expired_candidate_media() -> tuple[int, int]:
    cutoff = datetime.now(timezone.utc) - timedelta(days=settings.candidate_media_retention_days)
    rows = db.all(
        """SELECT id,exported_at FROM highlight_candidates
           WHERE status='exported' AND media_cleaned_at='' AND exported_at<>''"""
    )
    count = total = 0
    for row in rows:
        try:
            exported = datetime.fromisoformat(str(row["exported_at"]).replace("Z", "+00:00"))
            if exported.tzinfo is None:
                exported = exported.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
        if exported > cutoff:
            continue
        result = cleanup_candidate_media(int(row["id"]))
        count += 1
        total += int(result["released_bytes"])
    if count:
        db.event("info", "candidate_retention_cleanup", f"已清理 {count} 条导出超过保留期的本机成片", {
            "count": count, "released_bytes": total,
            "retention_days": settings.candidate_media_retention_days,
        })
    return count, total


def _database_time(value: str) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed.replace(tzinfo=timezone.utc) if parsed.tzinfo is None else parsed.astimezone(timezone.utc)
    except ValueError:
        return None


def _elapsed_text(seconds: float) -> str:
    seconds = max(0, int(seconds))
    if seconds < 60:
        return f"{seconds}秒"
    if seconds < 3600:
        return f"{seconds // 60}分{seconds % 60:02d}秒"
    return f"{seconds // 3600}小时{(seconds % 3600) // 60:02d}分"


def _recent_time_text(value: str) -> tuple[str, float | None]:
    parsed = _database_time(value)
    if parsed is None:
        return "暂无完成记录", None
    elapsed = max(0.0, (datetime.now(timezone.utc) - parsed).total_seconds())
    local = parsed.astimezone(timezone(timedelta(hours=8))).strftime("%H:%M:%S")
    return f"{local}（{_elapsed_text(elapsed)}前）", elapsed


def processing_health(totals: dict[str, Any]) -> list[dict[str, Any]]:
    renderer = pipeline.media.renderer_status()
    ai_routes = pipeline.ai_route_status()
    ai_active = db.all(
        """SELECT s.id,s.source_id,s.status,s.updated_at,r.sequence,r.name AS room_name
           FROM recording_segments s LEFT JOIN live_rooms r ON r.id=s.room_id
           WHERE s.status IN ('gpt_analyzing','deepseek_analyzing')
           ORDER BY s.updated_at LIMIT ?""",
        (max(1, int(ai_routes["worker_count"])),),
    )
    render_active = db.all(
        """SELECT h.id,h.source_id,h.render_phase,h.render_started_at,h.render_worker,
                  h.render_encoder,h.updated_at,r.sequence,r.name AS room_name
           FROM highlight_candidates h LEFT JOIN live_rooms r ON r.id=h.room_id
           WHERE h.status='rendering' ORDER BY h.render_started_at,h.id LIMIT 2"""
    )
    stage_specs = [
        {
            "key": "asr", "title": "本机转写", "waiting": int(totals["waiting_asr"]),
            "warn": 45 * 60, "danger": 90 * 60,
            "active": db.one(
                """SELECT s.id,s.source_id,s.updated_at,r.sequence,r.name AS room_name
                   FROM recording_segments s LEFT JOIN live_rooms r ON r.id=s.room_id
                   WHERE s.status='transcribing' ORDER BY s.updated_at LIMIT 1"""
            ),
            "last": (db.one("SELECT MAX(transcribed_at) AS value FROM recording_segments") or {}).get("value", ""),
            "noun": "分片",
        },
        {
            "key": "ai", "title": f"多模型分析 · {ai_routes['worker_count']}路", "waiting": int(totals["waiting_ai"]),
            "warn": 10 * 60, "danger": 25 * 60,
            "active": ai_active[0] if ai_active else None,
            "active_items": ai_active,
            "last": (db.one("SELECT MAX(analyzed_at) AS value FROM recording_segments") or {}).get("value", ""),
            "noun": "分片",
        },
        {
            "key": "render", "title": f"视频渲染 · {renderer['mode']}", "waiting": int(totals["waiting_render"]),
            "warn": 10 * 60, "danger": 25 * 60,
            "active": render_active[0] if render_active else None,
            "active_items": render_active,
            "last": (db.one(
                """SELECT MAX(updated_at) AS value FROM highlight_candidates
                   WHERE preview_path<>'' AND status NOT IN ('visual_review','rendering')"""
            ) or {}).get("value", ""),
            "noun": "候选",
        },
    ]
    now = datetime.now(timezone.utc)
    result: list[dict[str, Any]] = []
    for spec in stage_specs:
        active = spec["active"]
        last_text, last_elapsed = _recent_time_text(str(spec["last"] or ""))
        item = {
            "key": spec["key"], "title": spec["title"], "waiting": spec["waiting"],
            "active": bool(active), "last_completed": last_text,
            "state": "idle", "state_label": "空闲", "current": "当前没有处理任务",
            "elapsed": "—", "hint": "队列为空，运行正常",
        }
        if spec["key"] == "ai" and ai_routes.get("circuit_open"):
            remaining = int(ai_routes.get("circuit_remaining_seconds") or 0)
            item.update(
                state="stalled", state_label="中转站保护中",
                current="检测到连续连接故障，已暂停所有模型请求，防止重复计费",
                elapsed=f"{_elapsed_text(remaining)}后重试",
                hint=str(ai_routes.get("last_error") or "等待中转站连接恢复")[:240],
            )
            result.append(item)
            continue
        if active:
            started = _database_time(str(active.get("render_started_at") or active.get("updated_at") or ""))
            elapsed = max(0.0, (now - started).total_seconds()) if started else 0.0
            room_label = " · ".join(part for part in (
                str(active.get("sequence") or ""), str(active.get("room_name") or active.get("source_id") or "")
            ) if part)
            item["current"] = f"{spec['noun']} #{active['id']}" + (f" · {room_label}" if room_label else "")
            if spec["key"] == "render":
                phase_labels = {
                    "preparing": "准备素材", "scanning_timeline": "扫描音画时间轴",
                    "timeline_cached": "复用时间轴缓存", "composing": "裁切拼接",
                    "timeline_fallback": "快速估算时间轴",
                    "encoding": "视频编码", "gpu_fallback": "显卡失败，切换 CPU",
                    "complete": "完成", "failed": "失败",
                }
                descriptions = []
                for job in spec.get("active_items", []):
                    label = phase_labels.get(str(job.get("render_phase") or ""), "准备素材")
                    room = " · ".join(part for part in (
                        str(job.get("sequence") or ""), str(job.get("room_name") or job.get("source_id") or "")
                    ) if part)
                    descriptions.append(
                        f"#{job['id']} {room} · {label} · {job.get('render_encoder') or renderer['encoder']}"
                    )
                item["current"] = "；".join(descriptions)
                item["hint"] = f"实际编码器 {renderer['encoder']}；{renderer['note']}"
            elif spec["key"] == "ai":
                descriptions = []
                for job in spec.get("active_items", []):
                    stage = "DeepSeek 补漏" if job.get("status") == "deepseek_analyzing" else "GPT 主选"
                    room = " · ".join(part for part in (
                        str(job.get("sequence") or ""), str(job.get("room_name") or job.get("source_id") or "")
                    ) if part)
                    descriptions.append(f"#{job['id']} {room} · {stage}")
                item["current"] = "；".join(descriptions)
                fallback = f"；备用 {ai_routes['fallback_model']}" if ai_routes["fallback_model"] else ""
                item["hint"] = f"主线路 {ai_routes['label']}{fallback}；失败会自动换线并保留任务"
            item["elapsed"] = _elapsed_text(elapsed)
            if elapsed >= spec["danger"]:
                item.update(state="stalled", state_label="可能卡住", hint="单项运行时间明显过长，建议查看任务管理器或错误日志")
            elif elapsed >= spec["warn"]:
                item.update(state="slow", state_label="处理较慢", hint="任务仍在运行，可继续观察CPU和队列变化")
            else:
                item.update(state="working", state_label="正常处理中", hint="检测到当前任务，处理线程正在工作")
                if spec["key"] == "render":
                    item["hint"] = f"实际编码器 {renderer['encoder']}；{renderer['note']}"
        elif spec["waiting"]:
            if spec["key"] == "ai" and int(ai_routes.get("delayed_count") or 0):
                delayed = int(ai_routes["delayed_count"])
                paused = int(ai_routes.get("paused_count") or 0)
                item.update(
                    state="waiting", state_label="退避等待",
                    current=f"{delayed} 个任务正在延迟重试" + (f"，{paused} 个已暂停" if paused else ""),
                    hint="故障任务不会立即循环；后续正常任务仍可继续处理",
                )
                result.append(item)
                continue
            if last_elapsed is not None and last_elapsed >= spec["danger"]:
                item.update(state="stalled", state_label="可能停滞", current="有任务积压，但当前未检测到处理项", hint="最近完成时间较久，建议刷新后继续观察或查看运行日志")
            else:
                item.update(state="waiting", state_label="等待调度", current="正在准备领取下一项", hint="队列存在任务，短暂没有当前项属于正常切换")
        result.append(item)
    return result


def handoff_segment_reports(job_ids: list[int]) -> list[dict[str, Any]]:
    placeholders = ",".join("?" for _ in job_ids)
    sessions = db.all(
        f"""SELECT DISTINCT h.session_id FROM publish_jobs p
            JOIN highlight_candidates h ON h.id=p.candidate_id
            WHERE p.id IN ({placeholders})""", job_ids,
    )
    reports: list[dict[str, Any]] = []
    seen: set[int] = set()
    for session in sessions:
        for segment in db.all(
            "SELECT id FROM recording_segments WHERE session_id=? ORDER BY id", (session["session_id"],)
        ):
            segment_id = int(segment["id"])
            if segment_id in seen:
                continue
            seen.add(segment_id)
            reports.append(segment_cleanup_report(segment_id))
    return reports


class WebUpdateRequest(BaseModel):
    confirm: bool = False


@app.get("/api/update/status")
def web_update_status() -> dict[str, Any]:
    return github_update_status()


@app.get("/api/update/current")
def web_update_current() -> dict[str, str]:
    return {"current_version": current_app_version()}


@app.post("/api/update/install")
def web_update_install(payload: WebUpdateRequest) -> dict[str, Any]:
    if not payload.confirm:
        raise HTTPException(422, "必须明确确认后才能安装更新")
    active_tasks = db.all(
        """SELECT id,status FROM recording_segments
           WHERE status IN ('transcribing','gpt_analyzing','deepseek_analyzing')
           UNION ALL
           SELECT id,status FROM highlight_candidates WHERE status='rendering'"""
    )
    active_rooms = db.all(
        "SELECT sequence,name FROM live_rooms WHERE archived=0 AND enabled=1 AND live_status='live' ORDER BY sequence"
    )
    blockers: list[str] = []
    if active_tasks:
        blockers.append(f"还有 {len(active_tasks)} 个转写、模型或渲染任务正在运行")
    if active_rooms:
        names = "、".join(f"{row['sequence']} {row['name']}" for row in active_rooms[:3])
        blockers.append(f"仍有直播间开启录制：{names}" + ("等" if len(active_rooms) > 3 else ""))
    if blockers:
        raise HTTPException(409, {"message": "当前不适合更新，请先暂停录制并等待正在处理的任务完成", "blockers": blockers})
    status = github_update_status()
    if not status.get("ok"):
        raise HTTPException(503, status.get("message") or "暂时无法检查更新")
    if not status.get("update_available"):
        return {"ok": True, "started": False, **status}
    root = settings.service_root.parent
    update_script = root / "update.ps1"
    if not update_script.is_file():
        raise HTTPException(409, "程序目录缺少 update.ps1，请先使用完整包或一次性更新补丁")
    creationflags = getattr(subprocess, "CREATE_NEW_CONSOLE", 0)
    subprocess.Popen(
        ["powershell.exe", "-NoLogo", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File",
         str(update_script), "-WebInstall", str(os.getpid())],
        cwd=str(root),
        creationflags=creationflags,
    )
    return {
        "ok": True,
        "started": True,
        "current_version": status["current_version"],
        "available_version": status["available_version"],
        "message": "安全更新程序已启动；完成后中控台会自动重启",
    }


@app.get("/", response_class=HTMLResponse)
def control_center(request: Request) -> HTMLResponse:
    cleanup_expired_candidate_media()
    cards = room_cards()
    totals = {
        "rooms": len(cards),
        "recording": sum(room["live_state"] == "live" for room in cards),
        "active_recordings": sum(bool(room["recording_active"]) for room in cards),
        "segments": sum(int(room["segment_total"]) for room in cards),
        "transcribed": sum(int(room["segment_transcribed"]) for room in cards),
        "model_submitted": sum(int(room["segment_model_submitted"]) for room in cards),
        "analyzed": sum(int(room["segment_analyzed"]) for room in cards),
        "yielded": sum(int(room["segment_yielded"]) for room in cards),
        "candidates": sum(int(room["candidate_total"]) for room in cards),
        "effective": sum(int(room["effective_count"]) for room in cards),
        "pending": sum(int(room["pending_review"]) for room in cards),
        "waiting_asr": sum(int(room["waiting_asr"]) for room in cards),
        "waiting_ai": sum(int(room["waiting_ai"]) for room in cards),
        "waiting_render": sum(int(room["waiting_render"]) for room in cards),
        "active_asr": sum(int(room["active_asr"]) for room in cards),
        "active_ai": sum(int(room["active_ai"]) for room in cards),
        "active_render": sum(int(room["active_render"]) for room in cards),
        "render_errors": sum(int(room["render_error_count"]) for room in cards),
    }
    recent_events = db.all(
        "SELECT level,event_type,message,created_at FROM service_events ORDER BY id DESC LIMIT 12"
    )
    return templates.TemplateResponse(request, "dashboard.html", {
        "rooms": cards, "totals": totals,
        "app_version": current_app_version(),
        "processing_health": processing_health(totals),
        "recorder_running": pipeline.recorder.running,
        "cloud_text": pipeline.analyzer.cloud.text_enabled,
        "ai_routes": pipeline.ai_route_status(),
        "deepseek_supplement": bool(pipeline.supplement_analyzer),
        "recent_events": recent_events,
    })


@app.get("/review", response_class=HTMLResponse)
def dashboard(request: Request, status: str = "pending_review", room_id: str = "",
              model: str = "", output_date: str = "") -> HTMLResponse:
    cleanup_expired_candidate_media()
    selected_room_id = optional_query_int(room_id, "直播间")
    conditions: list[str] = []
    params: list[Any] = []
    if status == "all":
        pass
    else:
        conditions.append("status=?"); params.append(status)
    if selected_room_id is not None:
        conditions.append("room_id=?"); params.append(selected_room_id)
    if model == "gpt":
        conditions.append("analysis_version NOT LIKE '%deepseek%'")
    elif model == "deepseek":
        conditions.append("analysis_version LIKE '%deepseek%'")
    date_range = candidate_output_date_range(output_date)
    if date_range:
        conditions.append("created_at>=? AND created_at<?")
        params.extend(date_range)
    where = " WHERE " + " AND ".join(conditions) if conditions else ""
    rows = db.all(f"SELECT * FROM highlight_candidates{where} ORDER BY created_at DESC LIMIT 200", params)
    segment_counts = db.all("SELECT status,COUNT(*) AS count FROM recording_segments GROUP BY status")
    events = db.all("SELECT * FROM service_events ORDER BY id DESC LIMIT 12")
    return templates.TemplateResponse(request, "index.html", {
        "candidates": [candidate_view(row) for row in rows],
        "segment_counts": segment_counts,
        "events": events,
        "selected_status": status,
        "selected_room_id": selected_room_id,
        "selected_model": model,
        "selected_output_date": output_date,
        "rooms": db.all("SELECT id,sequence,name FROM live_rooms WHERE archived=0 ORDER BY sequence"),
        "asr_available": pipeline.transcriber.available,
        "asr_runtime": pipeline.transcriber.runtime,
        "cloud_text": pipeline.analyzer.cloud.text_enabled,
        "ai_routes": pipeline.ai_route_status(),
        "deepseek_supplement": bool(pipeline.supplement_analyzer),
        "recorder_running": pipeline.recorder.running,
        "cloud_vision": pipeline.analyzer.cloud.vision_enabled,
        "input_dir": str(settings.input_dir),
    })


@app.get("/export", response_class=HTMLResponse)
def export_center(request: Request, status: str = "accepted", room_id: str = "", source_id: str = "",
                  model: str = "", output_date: str = "", export_date: str = "",
                  date_scope: str = "") -> HTMLResponse:
    cleanup_expired_candidate_media()
    selected_room_id = optional_query_int(room_id, "直播间")
    selected_source_id = source_id.strip()
    if status not in {"accepted", "exported", "all"}:
        status = "accepted"
    china_tz = timezone(timedelta(hours=8))
    today_label = datetime.now(china_tz).strftime("%Y-%m-%d")
    selected_date_scope = "all" if date_scope == "all" else "day"
    effective_output_date = "" if selected_date_scope == "all" else (output_date.strip() or today_label)
    conditions = ["h.status IN ('accepted','exported')"] if status == "all" else ["h.status=?"]
    params: list[Any] = [] if status == "all" else [status]
    if selected_room_id is not None:
        conditions.append("h.room_id=?")
        params.append(selected_room_id)
    elif selected_source_id:
        conditions.append("h.room_id IS NULL AND h.source_id=?")
        params.append(selected_source_id)
    if model == "gpt":
        conditions.append("h.analysis_version NOT LIKE '%deepseek%'")
    elif model == "deepseek":
        conditions.append("h.analysis_version LIKE '%deepseek%'")
    output_range = candidate_output_date_range(effective_output_date)
    if output_range:
        conditions.append("h.created_at>=? AND h.created_at<?")
        params.extend(output_range)
    exported_range = candidate_output_date_range(export_date)
    if exported_range:
        conditions.append("h.exported_at>=? AND h.exported_at<?")
        params.extend(exported_range)
    rows = db.all(
        "SELECT h.*,r.sequence AS room_sequence,r.name AS room_name FROM highlight_candidates h "
        "LEFT JOIN live_rooms r ON r.id=h.room_id WHERE " + " AND ".join(conditions)
        + " ORDER BY CASE WHEN h.exported_at='' THEN h.created_at ELSE h.exported_at END DESC LIMIT 500",
        params,
    )
    card_conditions = ["h.status IN ('accepted','exported')"]
    card_params: list[Any] = []
    if output_range:
        card_conditions.append("h.created_at>=? AND h.created_at<?")
        card_params.extend(output_range)
    if model == "gpt":
        card_conditions.append("h.analysis_version NOT LIKE '%deepseek%'")
    elif model == "deepseek":
        card_conditions.append("h.analysis_version LIKE '%deepseek%'")
    room_export_cards = db.all(
        """SELECT h.room_id,h.source_id,r.sequence,r.name AS room_name,
                  COUNT(*) AS produced,
                  SUM(h.status='exported') AS exported,
                  SUM(h.status='accepted') AS not_exported
           FROM highlight_candidates h
           LEFT JOIN live_rooms r ON r.id=h.room_id
           WHERE """ + " AND ".join(card_conditions) + """
           GROUP BY CASE WHEN h.room_id IS NOT NULL THEN 'room:' || h.room_id ELSE 'source:' || h.source_id END
           ORDER BY COALESCE(r.sequence,'999999'),r.name,h.source_id""",
        card_params,
    )
    date_totals = {
        "produced": sum(int(row["produced"] or 0) for row in room_export_cards),
        "exported": sum(int(row["exported"] or 0) for row in room_export_cards),
        "not_exported": sum(int(row["not_exported"] or 0) for row in room_export_cards),
    }
    selected_room_label = ""
    if selected_room_id is not None:
        selected = next((row for row in room_export_cards if row.get("room_id") == selected_room_id), None)
        if selected:
            selected_room_label = f"{selected.get('sequence') or ''} · {selected.get('room_name') or selected.get('source_id') or '未知直播间'}".strip(" ·")
    elif selected_source_id:
        selected_room_label = selected_source_id
    return templates.TemplateResponse(request, "export.html", {
        "candidates": [candidate_view(row) for row in rows],
        "selected_status": status,
        "selected_room_id": selected_room_id,
        "selected_source_id": selected_source_id,
        "selected_room_label": selected_room_label,
        "selected_model": model,
        "selected_output_date": effective_output_date,
        "selected_export_date": export_date,
        "selected_date_scope": selected_date_scope,
        "rooms": db.all("SELECT id,sequence,name FROM live_rooms WHERE archived=0 ORDER BY sequence"),
        "waiting_export": date_totals["not_exported"],
        "exported_count": date_totals["exported"],
        "today_label": today_label,
        "date_label": effective_output_date or "全部日期",
        "room_export_cards": room_export_cards,
        "date_totals": date_totals,
        "show_materials": selected_room_id is not None or bool(selected_source_id),
        "retention_days": settings.candidate_media_retention_days,
    })


@app.get("/catalog", response_class=HTMLResponse)
def catalog_page(request: Request, q: str = "") -> HTMLResponse:
    return RedirectResponse("/", status_code=302)
    pattern = f"%{q.strip()}%"
    items = db.all(
        """SELECT i.*,
           (SELECT COUNT(*) FROM live_rooms r WHERE r.default_catalog_item_id=i.id) AS room_uses,
           (SELECT COUNT(*) FROM recording_segments s WHERE s.catalog_item_id=i.id) AS segment_uses,
           (SELECT COUNT(*) FROM highlight_candidates h WHERE h.catalog_item_id=i.id) AS candidate_uses,
           (SELECT COUNT(*) FROM publish_jobs p WHERE p.catalog_item_id=i.id) AS handoff_uses,
           (SELECT COUNT(*) FROM review_decisions d WHERE d.catalog_item_id=i.id) AS review_uses
           FROM catalog_items i
           WHERE i.active=1 AND (?='' OR i.internal_code LIKE ? OR i.name LIKE ? OR i.aliases LIKE ?)
           ORDER BY i.internal_code LIMIT 1000""",
        (q.strip(), pattern, pattern, pattern),
    )
    return templates.TemplateResponse(request, "catalog.html", {"items": items, "q": q})


@app.get("/api/health")
def health() -> dict[str, Any]:
    return {
        "ok": True,
        "queue_size": pipeline.queue.qsize(),
        "stage_queues": {
            "waiting_asr": db.one("SELECT COUNT(id) count FROM recording_segments WHERE status='discovered'")["count"],
            "waiting_ai": db.one("SELECT COUNT(id) count FROM recording_segments WHERE status IN ('transcribed','ai_waiting')")["count"],
            "waiting_render": db.one(
                """SELECT COUNT(h.id) count FROM highlight_candidates h WHERE h.status='visual_review'
                   AND EXISTS (SELECT 1 FROM recording_segments s WHERE s.session_id=h.session_id
                     AND NOT(s.timeline_end<=h.start_time OR s.timeline_start>=h.end_time))
                   AND NOT EXISTS (SELECT 1 FROM recording_segments s WHERE s.session_id=h.session_id
                     AND NOT(s.timeline_end<=h.start_time OR s.timeline_start>=h.end_time)
                     AND s.status NOT IN ('complete','analyzed'))"""
            )["count"],
        },
        "asr_available": pipeline.transcriber.available,
        "cloud_text_enabled": pipeline.analyzer.cloud.text_enabled,
        "ai_routes": pipeline.ai_route_status(),
        "deepseek_supplement_enabled": bool(pipeline.supplement_analyzer),
        "recorder_running": pipeline.recorder.running,
        "cloud_vision_enabled": pipeline.analyzer.cloud.vision_enabled,
        "settings": settings.public_dict(),
    }


@app.post("/api/config/open-model-keys")
def open_model_key_config() -> dict[str, Any]:
    """Add missing multi-model fields without changing existing secrets, then open Notepad."""
    env_path = settings.service_root / ".env"
    existing = env_path.read_text(encoding="utf-8") if env_path.exists() else ""
    present = {
        line.split("=", 1)[0].strip()
        for line in existing.splitlines()
        if "=" in line and not line.lstrip().startswith("#")
    }
    defaults = [
        ("HIGHLIGHT_AI_BASE_URL", "https://api.sisct2.xyz/v1"),
        ("HIGHLIGHT_AI_API_KEY", ""),
        ("HIGHLIGHT_AI_MODEL", "gpt-5.5"),
        ("HIGHLIGHT_AI_SECONDARY_API_KEY", ""),
        ("HIGHLIGHT_AI_SECONDARY_MODEL", "gpt-5.4"),
        ("HIGHLIGHT_AI_FALLBACK_API_KEY", ""),
        ("HIGHLIGHT_AI_FALLBACK_MODEL", "gpt-5.4-mini"),
        ("HIGHLIGHT_AI_WORKER_COUNT", "2"),
        ("HIGHLIGHT_GUARDIAN_AI_BASE_URL", "https://api.sisct2.xyz/v1"),
        ("HIGHLIGHT_GUARDIAN_AI_API_KEY", ""),
        ("HIGHLIGHT_GUARDIAN_AI_MODEL", "gpt-5.5"),
        ("HIGHLIGHT_GUARDIAN_VISION_API_KEY", ""),
        ("HIGHLIGHT_GUARDIAN_VISION_MODEL", "gpt-5.5"),
    ]
    missing = [(key, value) for key, value in defaults if key not in present]
    if missing:
        prefix = "\n" if existing and not existing.endswith("\n") else ""
        addition = prefix + "\n# GPT 多模型并行与自动备用线路\n" + "".join(
            f"{key}={value}\n" for key, value in missing
        )
        env_path.parent.mkdir(parents=True, exist_ok=True)
        with env_path.open("a", encoding="utf-8") as handle:
            handle.write(addition)
    creationflags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
    subprocess.Popen(["notepad.exe", str(env_path)], creationflags=creationflags)
    return {
        "ok": True,
        "message": "密钥配置已打开；保存后请关闭中控台并重新启动，配置才会生效",
        "added_fields": [key for key, _ in missing],
    }


class GuardianActionRequest(BaseModel):
    action: Literal["abandon", "resume"]
    segment_ids: list[int] = Field(default_factory=list, max_length=200)
    confirm: bool = False


def guardian_problem_ids(include_abandoned: bool = False, message: str = "") -> list[int]:
    statuses = "('ai_waiting','ai_retry_paused','ai_abandoned')" if include_abandoned else "('ai_waiting','ai_retry_paused')"
    rows = db.all(
        f"""SELECT s.id,COALESCE(r.sequence,'') AS room_sequence,COALESCE(r.name,'') AS room_name
            FROM recording_segments s LEFT JOIN live_rooms r ON r.id=s.room_id
            WHERE s.status IN {statuses} ORDER BY s.id LIMIT 200"""
    )
    requested_ids = {
        int(value) for value in re.findall(r"(?:任务|分片|#)\s*#?(\d+)", message, flags=re.I)
    }
    requested_rooms = set(re.findall(r"直播间\s*0*(\d{1,6})", message))
    named_rooms = {
        str(row["room_name"]) for row in rows
        if len(str(row.get("room_name") or "")) >= 2 and str(row["room_name"]) in message
    }
    if requested_ids:
        rows = [row for row in rows if int(row["id"]) in requested_ids]
    elif requested_rooms or named_rooms:
        rows = [
            row for row in rows
            if str(row.get("room_sequence") or "").lstrip("0") in requested_rooms
            or str(row.get("room_name") or "") in named_rooms
        ]
    return [int(row["id"]) for row in rows]


@app.post("/api/guardian/chat")
async def guardian_chat(message: str = Form(default=""), image: UploadFile | None = File(default=None)) -> dict[str, Any]:
    message = message.strip() or ("请分析这张程序截图" if image else "播报当前程序状态")
    lowered = message.lower()
    if "导出" in message and any(word in message for word in ("诊断", "日志", "数据包")):
        package = guardian.export_diagnostics()
        return {
            "ok": True, "source": "local", "severity": "ok",
            "answer": "脱敏诊断包已经生成，不包含密钥、Cookie、转写正文、录像或完整数据库。",
            "download_url": f"/api/guardian/diagnostics/{quote(package.name)}",
        }
    if any(word in message for word in ("舍弃", "放弃", "不要这批")):
        ids = guardian_problem_ids(message=message)
        if not ids:
            return {"ok": True, "source": "local", "severity": "ok", "answer": "当前没有可舍弃的异常模型任务。"}
        return {
            "ok": True, "source": "local", "severity": "warning",
            "answer": f"找到 {len(ids)} 个正在重试或已暂停的模型任务。舍弃只会停止这些任务，不会立即删除录像和转写；以后仍可恢复。",
            "pending_action": {"action": "abandon", "segment_ids": ids, "label": f"确认舍弃 {len(ids)} 个任务"},
        }
    if any(word in message for word in ("继续尝试", "继续重试", "恢复任务", "重新尝试")):
        ids = guardian_problem_ids(include_abandoned=True, message=message)
        if not ids:
            return {"ok": True, "source": "local", "severity": "ok", "answer": "当前没有需要恢复的模型任务。"}
        return {
            "ok": True, "source": "local", "severity": "warning",
            "answer": f"可以恢复 {len(ids)} 个延迟、暂停或已舍弃的任务，同时解除中转站保护。恢复后会重新产生模型请求和费用。",
            "pending_action": {"action": "resume", "segment_ids": ids, "label": f"确认恢复 {len(ids)} 个任务"},
        }
    image_payload: tuple[bytes, str] | None = None
    if image:
        mime = str(image.content_type or "")
        if mime not in {"image/png", "image/jpeg", "image/webp"}:
            raise HTTPException(422, "只支持 PNG、JPG 或 WebP 截图")
        raw = await image.read(5 * 1024 * 1024 + 1)
        if len(raw) > 5 * 1024 * 1024:
            raise HTTPException(413, "截图不能超过 5MB")
        image_payload = (raw, mime)
    return {"ok": True, **guardian.answer(message, image=image_payload)}


@app.post("/api/guardian/action")
def guardian_action(payload: GuardianActionRequest) -> dict[str, Any]:
    if not payload.confirm:
        raise HTTPException(422, "必须在网页中明确确认后才能执行")
    ids = sorted({int(item) for item in payload.segment_ids if int(item) > 0})
    if not ids:
        raise HTTPException(422, "没有选择任务")
    if payload.action == "abandon":
        changed = pipeline.abandon_ai_segments(ids)
        db.event("warning", "guardian_abandon", f"用户通过AI管家舍弃了 {changed} 个异常模型任务", {"segment_ids": ids})
        return {"ok": True, "message": f"已停止 {changed} 个任务；数据仍保留，可通过管家恢复"}
    changed = pipeline.resume_ai_segments(ids)
    pipeline.reset_ai_circuit(clear_disabled=True)
    db.event("info", "guardian_resume", f"用户通过AI管家恢复了 {changed} 个模型任务", {"segment_ids": ids})
    return {"ok": True, "message": f"已恢复 {changed} 个任务，并解除中转站保护"}


@app.get("/api/guardian/diagnostics/{filename}")
def guardian_diagnostics_download(filename: str) -> FileResponse:
    safe_name = Path(filename).name
    if safe_name != filename or not safe_name.startswith("AI管家诊断包_") or not safe_name.endswith(".zip"):
        raise HTTPException(404, "诊断包不存在")
    path = settings.data_dir / "diagnostics" / safe_name
    if not path.is_file():
        raise HTTPException(404, "诊断包不存在")
    return FileResponse(path, filename=safe_name, media_type="application/zip")


class RoomRequest(BaseModel):
    sequence: str = Field(default="", max_length=20)
    name: str = Field(min_length=1, max_length=200)
    url: str = Field(min_length=8, max_length=1000)
    enabled: bool = True
    notes: str = ""


class RoomPatch(BaseModel):
    sequence: str | None = Field(default=None, min_length=1, max_length=20)
    name: str | None = Field(default=None, min_length=1, max_length=200)
    url: str | None = Field(default=None, min_length=8, max_length=1000)
    enabled: bool | None = None
    archived: bool | None = None
    notes: str | None = None


class RoomToggle(BaseModel):
    enabled: bool


class RoomBatchRequest(BaseModel):
    text: str = Field(min_length=1, max_length=100000)
    enabled: bool = True


def normalize_sequence(value: str) -> str:
    value = value.strip()
    return value.zfill(3) if value.isdigit() else value


def next_room_sequence(database: Database | None = None) -> str:
    target = database or db
    numeric = [int(row["sequence"]) for row in target.all("SELECT sequence FROM live_rooms")
               if str(row["sequence"]).isdigit()]
    return f"{(max(numeric, default=0) + 1):03d}"


def validate_room_url(value: str) -> str:
    value = value.strip()
    parsed = urlparse(value)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise HTTPException(422, "请输入有效的直播间链接")
    return value


def canonical_room_url(value: str) -> str:
    parsed_item = RoomRegistry.parse_line(value)
    normalized = validate_room_url(parsed_item["url"] if parsed_item else value)
    parsed = urlparse(normalized)
    path = parsed.path.rstrip("/") or "/"
    return f"{parsed.scheme.lower()}://{parsed.netloc.lower()}{path}"


def analyze_room_batch(text: str, database: Database | None = None) -> dict[str, Any]:
    target = database or db
    lines = [line.strip() for line in text.replace("\ufeff", "").splitlines() if line.strip()]
    existing_rooms = target.all("SELECT id,sequence,name,url,source_key,archived FROM live_rooms ORDER BY id")
    by_url: dict[str, dict[str, Any]] = {}
    by_key: dict[str, dict[str, Any]] = {}
    for room in existing_rooms:
        try:
            by_url[canonical_room_url(room["url"])] = room
        except HTTPException:
            pass
        by_key[str(room["source_key"])] = room

    items: list[dict[str, Any]] = []
    pending_urls: set[str] = set()
    pending_keys: set[str] = set()
    pair_count = (len(lines) + 1) // 2
    for index in range(pair_count):
        name = lines[index * 2].strip()
        raw_url = lines[index * 2 + 1].strip() if index * 2 + 1 < len(lines) else ""
        item: dict[str, Any] = {"index": index + 1, "name": name, "url": raw_url}
        try:
            if not raw_url:
                raise ValueError("缺少链接行")
            if name.lower().startswith(("http://", "https://")):
                raise ValueError("第一行应为直播间名称，第二行才是链接")
            parsed_item = RoomRegistry.parse_line(raw_url)
            normalized_url = validate_room_url(parsed_item["url"] if parsed_item else raw_url)
            canonical_url = canonical_room_url(normalized_url)
            key = safe_id(name)
            item["url"] = normalized_url
            existing = by_url.get(canonical_url) or by_key.get(key)
            if existing:
                item.update({
                    "status": "existing", "status_label": "已存在，跳过",
                    "reason": f"已收录为 {existing['sequence']} · {existing['name']}"
                              + ("（已归档）" if existing.get("archived") else ""),
                })
            elif canonical_url in pending_urls or key in pending_keys:
                item.update({
                    "status": "duplicate", "status_label": "本次重复，跳过",
                    "reason": "同一批列表里已经出现过相同名称或链接",
                })
            else:
                item.update({"status": "new", "status_label": "未收录，可添加", "reason": ""})
                pending_urls.add(canonical_url)
                pending_keys.add(key)
        except (HTTPException, ValueError, KeyError) as exc:
            detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
            item.update({"status": "invalid", "status_label": "格式有误", "reason": detail})
        items.append(item)

    return {
        "items": items,
        "total": len(items),
        "new_count": sum(item["status"] == "new" for item in items),
        "existing_count": sum(item["status"] == "existing" for item in items),
        "duplicate_count": sum(item["status"] == "duplicate" for item in items),
        "invalid_count": sum(item["status"] == "invalid" for item in items),
    }


@app.get("/api/rooms")
def api_rooms() -> dict[str, Any]:
    return {"rooms": room_cards()}


@app.get("/api/rooms/export")
def export_rooms() -> Response:
    payload = export_room_backup(db)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        content=json.dumps(payload, ensure_ascii=False, indent=2),
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="rooms-backup-{stamp}.json"'},
    )


@app.post("/api/rooms/import")
def import_rooms(payload: dict[str, Any]) -> dict[str, Any]:
    try:
        result = import_room_backup(db, pipeline.rooms, payload)
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    pipeline.recorder.ensure_running()
    pipeline.live_monitor.request_refresh()
    db.event("info", "room_import", f"已导入直播间配置：新增 {result['created']}，更新 {result['updated']}")
    return {"ok": True, **result}


@app.post("/api/rooms/batch-check")
def check_room_batch(payload: RoomBatchRequest) -> dict[str, Any]:
    return {"ok": True, **analyze_room_batch(payload.text)}


@app.post("/api/rooms/batch-add")
def add_room_batch(payload: RoomBatchRequest) -> dict[str, Any]:
    analysis = analyze_room_batch(payload.text)
    new_items = [item for item in analysis["items"] if item["status"] == "new"]
    created_ids: list[int] = []
    created_rooms: list[dict[str, Any]] = []
    try:
        for item in new_items:
            sequence = next_room_sequence()
            now = utc_now()
            room_id = db.execute(
                """INSERT INTO live_rooms
                   (sequence,name,url,source_key,enabled,notes,created_at,updated_at)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (sequence, item["name"], item["url"], safe_id(item["name"]),
                 int(payload.enabled), "", now, now),
            )
            created_ids.append(room_id)
            created_rooms.append({"id": room_id, "sequence": sequence, "name": item["name"], "url": item["url"]})
        if created_ids:
            pipeline.rooms.sync()
    except Exception as exc:
        for room_id in reversed(created_ids):
            db.execute("DELETE FROM live_rooms WHERE id=?", (room_id,))
        try:
            pipeline.rooms.sync()
        except Exception:
            pass
        raise HTTPException(500, f"批量添加失败，未保留本次新增内容：{exc}") from exc

    if created_ids:
        pipeline.recorder.ensure_running()
        pipeline.live_monitor.request_refresh()
        db.event("info", "room_batch", f"批量核对后新增 {len(created_ids)} 个直播间", {
            "room_ids": created_ids,
        })
    return {"ok": True, "created": len(created_ids), "rooms": created_rooms, **analysis}


@app.post("/api/rooms")
def create_room(payload: RoomRequest) -> dict[str, Any]:
    sequence = normalize_sequence(payload.sequence) if payload.sequence.strip() else next_room_sequence()
    name = payload.name.strip()
    url = validate_room_url(payload.url)
    if db.one("SELECT id FROM live_rooms WHERE sequence=? OR url=?", (sequence, url)):
        raise HTTPException(409, "直播间序号或链接已经存在")
    key = safe_id(name)
    if db.one("SELECT id FROM live_rooms WHERE source_key=?", (key,)):
        raise HTTPException(409, "直播间名称生成的录制标识已经存在")
    now = utc_now()
    room_id = db.execute(
        """INSERT INTO live_rooms
           (sequence,name,url,source_key,enabled,notes,created_at,updated_at)
           VALUES(?,?,?,?,?,?,?,?)""",
        (sequence, name, url, key, int(payload.enabled), payload.notes.strip(), now, now),
    )
    try:
        pipeline.rooms.sync()
    except Exception as exc:
        db.execute("DELETE FROM live_rooms WHERE id=?", (room_id,))
        raise HTTPException(500, f"同步录制器配置失败：{exc}") from exc
    # A fresh installation has no room, so the recorder is deliberately not
    # started at service boot.  Saving the first enabled room starts it only
    # after URL_config.ini has been written.
    pipeline.recorder.ensure_running()
    pipeline.live_monitor.request_refresh()
    db.event("info", "room", f"新增直播间 {sequence} · {name}", {"room_id": room_id})
    return {"ok": True, "room": db.one("SELECT * FROM live_rooms WHERE id=?", (room_id,))}


@app.patch("/api/rooms/{room_id}")
def update_room(room_id: int, payload: RoomPatch) -> dict[str, Any]:
    room = db.one("SELECT * FROM live_rooms WHERE id=?", (room_id,))
    if not room:
        raise HTTPException(404, "直播间不存在")
    sequence = normalize_sequence(payload.sequence) if payload.sequence is not None else room["sequence"]
    name = payload.name.strip() if payload.name is not None else room["name"]
    url = validate_room_url(payload.url) if payload.url is not None else room["url"]
    duplicate = db.one("SELECT id FROM live_rooms WHERE id<>? AND (sequence=? OR url=?)", (room_id, sequence, url))
    if duplicate:
        raise HTTPException(409, "直播间序号或链接已经存在")
    key = safe_id(name)
    key_duplicate = db.one("SELECT id FROM live_rooms WHERE id<>? AND source_key=?", (room_id, key))
    if key_duplicate:
        raise HTTPException(409, "直播间名称生成的录制标识已经存在")
    values = (
        sequence, name, url, key,
        int(payload.enabled) if payload.enabled is not None else room["enabled"],
        int(payload.archived) if payload.archived is not None else room["archived"],
        payload.notes.strip() if payload.notes is not None else room["notes"], utc_now(), room_id,
    )
    db.execute(
        """UPDATE live_rooms SET sequence=?,name=?,url=?,source_key=?,enabled=?,archived=?,
           default_catalog_item_id=NULL,notes=?,updated_at=? WHERE id=?""", values,
    )
    try:
        pipeline.rooms.sync()
    except Exception as exc:
        db.execute(
            """UPDATE live_rooms SET sequence=?,name=?,url=?,source_key=?,enabled=?,archived=?,
               default_catalog_item_id=?,notes=?,updated_at=? WHERE id=?""",
            (room["sequence"], room["name"], room["url"], room["source_key"], room["enabled"],
             room["archived"], room.get("default_catalog_item_id"), room["notes"], room["updated_at"], room_id),
        )
        raise HTTPException(500, f"同步录制器配置失败：{exc}") from exc
    pipeline.live_monitor.request_refresh()
    return {"ok": True, "room": db.one("SELECT * FROM live_rooms WHERE id=?", (room_id,))}


@app.post("/api/rooms/{room_id}/toggle")
def toggle_room(room_id: int, payload: RoomToggle) -> dict[str, Any]:
    if not db.one("SELECT id FROM live_rooms WHERE id=?", (room_id,)):
        raise HTTPException(404, "直播间不存在")
    db.execute("UPDATE live_rooms SET enabled=?,updated_at=? WHERE id=?", (int(payload.enabled), utc_now(), room_id))
    pipeline.rooms.sync()
    pipeline.live_monitor.request_refresh()
    return {"ok": True, "enabled": payload.enabled}


@app.delete("/api/rooms/{room_id}")
def delete_room(room_id: int) -> dict[str, Any]:
    room = db.one("SELECT * FROM live_rooms WHERE id=?", (room_id,))
    if not room:
        raise HTTPException(404, "直播间不存在")
    if active_recording_state(room)["recording_active"]:
        raise HTTPException(409, "该直播间正在录制，请先暂停并等待录制文件封口后再删除")
    history = {
        "录像分片": int(db.one("SELECT COUNT(*) AS count FROM recording_segments WHERE room_id=?", (room_id,))["count"]),
        "候选素材": int(db.one("SELECT COUNT(*) AS count FROM highlight_candidates WHERE room_id=?", (room_id,))["count"]),
        "旧交接记录": int(db.one("SELECT COUNT(*) AS count FROM publish_jobs WHERE room_id=?", (room_id,))["count"]),
    }
    used = {label: count for label, count in history.items() if count}
    if used:
        summary = "、".join(f"{label}{count}条" for label, count in used.items())
        raise HTTPException(409, f"该直播间已有{summary}，不能直接删除；请使用归档以保留历史记录")
    db.execute("DELETE FROM live_rooms WHERE id=?", (room_id,))
    pipeline.rooms.sync()
    pipeline.live_monitor.request_refresh()
    db.event("info", "room", f"已删除误填直播间 {room['sequence']} · {room['name']}")
    return {"ok": True}


@app.get("/rooms/{room_id}", response_class=HTMLResponse)
def room_page(request: Request, room_id: int) -> HTMLResponse:
    room = next((item for item in room_cards() if item["id"] == room_id), None)
    if not room:
        raise HTTPException(404, "直播间不存在")
    segments = db.all("SELECT * FROM recording_segments WHERE room_id=? ORDER BY id DESC LIMIT 100", (room_id,))
    for segment in segments:
        segment["filename"] = Path(segment["path"]).name
        segment["stage_label"] = {
            "discovered": "等待处理", "transcribing": "转写中", "transcribed": "等待模型分析",
            "gpt_analyzing": "GPT分析中", "deepseek_analyzing": "DeepSeek补漏中",
            "ai_waiting": "模型重试中", "analyzed": "分析完成", "complete": "流程完成",
            "error": "处理异常", "asr_unavailable": "转写不可用",
        }.get(segment["status"], segment["status"])
    candidates = db.all("SELECT * FROM highlight_candidates WHERE room_id=? ORDER BY id DESC LIMIT 100", (room_id,))
    for segment in segments:
        segment["candidate_count"] = 0
        segment["gpt_candidate_count"] = 0
        segment["deepseek_candidate_count"] = 0
        segment["candidate_ids"] = []
    for candidate in candidates:
        ranges = json_field(candidate, "source_ranges_json", []) or [{
            "start": candidate["start_time"], "end": candidate["end_time"],
        }]
        sources = []
        for segment in segments:
            if segment["session_id"] != candidate["session_id"]:
                continue
            if not any(
                float(item["end"]) > float(segment["timeline_start"])
                and float(item["start"]) < float(segment["timeline_end"])
                for item in ranges
            ):
                continue
            sources.append({"id": segment["id"], "filename": segment["filename"]})
            if candidate["status"] not in {"superseded", "render_error"}:
                segment["candidate_count"] += 1
                segment["candidate_ids"].append(candidate["id"])
                if "deepseek" in str(candidate.get("analysis_version") or "").lower():
                    segment["deepseek_candidate_count"] += 1
                else:
                    segment["gpt_candidate_count"] += 1
        candidate["source_segments"] = sources
    for segment in segments:
        cleanup = segment_cleanup_report(int(segment["id"]))
        segment["cleanup_cleanable"] = cleanup["cleanable"]
        segment["cleanup_already_cleaned"] = cleanup["already_cleaned"]
        segment["cleanup_blockers"] = cleanup["blockers"]
        segment["cleanup_estimated_gb"] = cleanup["estimated_gb"]
    return templates.TemplateResponse(request, "room.html", {
        "room": room, "segments": segments, "candidates": candidates,
    })


@app.get("/api/rooms/{room_id}")
def room_detail(room_id: int) -> dict[str, Any]:
    room = next((item for item in room_cards() if item["id"] == room_id), None)
    if not room:
        raise HTTPException(404, "直播间不存在")
    return {
        "room": room,
        "segments": db.all("SELECT * FROM recording_segments WHERE room_id=? ORDER BY id DESC", (room_id,)),
        "candidates": db.all("SELECT * FROM highlight_candidates WHERE room_id=? ORDER BY id DESC", (room_id,)),
    }


@app.get("/api/segments/{segment_id}/cleanup-check")
def cleanup_check(segment_id: int) -> dict[str, Any]:
    return segment_cleanup_report(segment_id)


class SegmentCleanupRequest(BaseModel):
    confirm: bool = False


@app.post("/api/segments/{segment_id}/cleanup")
def cleanup_segment(segment_id: int, payload: SegmentCleanupRequest) -> dict[str, Any]:
    report = segment_cleanup_report(segment_id)
    if not payload.confirm:
        raise HTTPException(422, "必须明确确认后才能清理")
    if report["already_cleaned"]:
        return {"ok": True, "idempotent": True, "released_bytes": report["segment"].get("released_bytes", 0)}
    if not report["cleanable"]:
        raise HTTPException(409, {"message": "该分片尚不能安全清理", "blockers": report["blockers"]})

    segment = report["segment"]
    candidates = [
        candidate for candidate in db.all(
            "SELECT * FROM highlight_candidates WHERE session_id=?", (segment["session_id"],)
        ) if _candidate_overlaps_segment(candidate, segment)
        and candidate["status"] != "superseded"
    ]
    input_root = settings.input_dir.resolve()
    data_roots = [settings.output_dir.resolve(), settings.cache_dir.resolve(), settings.keyframe_dir.resolve()]
    released = media.invalidate_timeline(Path(segment["path"]))
    released += _unlink_cleanup_file(Path(segment["path"]), [input_root])
    released += _unlink_cleanup_file(settings.cache_dir / f"segment_{segment_id}.wav", data_roots)
    candidate_ids: list[int] = []
    cleared_candidate_ids: list[int] = []
    seen_paths: set[str] = set()
    for candidate in candidates:
        candidate_ids.append(int(candidate["id"]))
        # 已交接成片单独保留 7 天，分片清理只移除原始录像和可再生缓存。
        if candidate["status"] != "exported":
            cleared_candidate_ids.append(int(candidate["id"]))
            for field in ("preview_path", "output_path"):
                path_text = str(candidate.get(field) or "")
                if path_text and path_text not in seen_paths:
                    released += _unlink_cleanup_file(Path(path_text), data_roots)
                    seen_paths.add(path_text)
        for path in settings.keyframe_dir.glob(f"candidate_{candidate['id']}_*"):
            released += _unlink_cleanup_file(path, data_roots)
        for pattern in (f"candidate_{candidate['id']}.*", f"candidate_{candidate['id']}_*"):
            for path in settings.cache_dir.glob(pattern):
                released += _unlink_cleanup_file(path, data_roots)
    now = utc_now()
    if cleared_candidate_ids:
        placeholders = ",".join("?" for _ in cleared_candidate_ids)
        db.execute(
            f"UPDATE highlight_candidates SET preview_path='',output_path='',updated_at=? WHERE id IN ({placeholders})",
            (now, *cleared_candidate_ids),
        )
    db.execute(
        "UPDATE recording_segments SET status='cleaned',cleaned_at=?,released_bytes=?,updated_at=? WHERE id=?",
        (now, released, now, segment_id),
    )
    db.event("info", "segment_cleanup", f"已安全清理录像分片 #{segment_id}", {
        "segment_id": segment_id, "candidate_ids": candidate_ids, "released_bytes": released,
    })
    return {"ok": True, "segment_id": segment_id, "released_bytes": released,
            "released_gb": round(released / (1024 ** 3), 3), "candidate_ids": candidate_ids}


def cleanup_ready_segments(segment_ids: list[int]) -> dict[str, Any]:
    """Safely remove source media only after every related candidate has a final disposition."""
    cleaned: list[dict[str, Any]] = []
    blocked: list[dict[str, Any]] = []
    for segment_id in list(dict.fromkeys(segment_ids)):
        report = segment_cleanup_report(segment_id)
        if report["already_cleaned"]:
            continue
        if not report["cleanable"]:
            blocked.append({"segment_id": segment_id, "blockers": report["blockers"]})
            continue
        cleaned.append(cleanup_segment(segment_id, SegmentCleanupRequest(confirm=True)))
    released = sum(int(item["released_bytes"]) for item in cleaned)
    if cleaned:
        db.event("info", "automatic_segment_cleanup", f"已自动安全清理 {len(cleaned)} 个完成分片", {
            "segment_ids": [item["segment_id"] for item in cleaned],
            "released_bytes": released,
            "rule": "all_candidates_exported_or_rejected",
        })
    return {
        "cleaned_count": len(cleaned),
        "cleaned_segment_ids": [item["segment_id"] for item in cleaned],
        "released_bytes": released,
        "released_gb": round(released / (1024 ** 3), 3),
        "blocked": blocked,
    }


def cleanup_ready_segments_for_candidate(candidate_id: int) -> dict[str, Any]:
    candidate = db.one("SELECT * FROM highlight_candidates WHERE id=?", (candidate_id,))
    if not candidate:
        return {"cleaned_count": 0, "cleaned_segment_ids": [], "released_bytes": 0, "released_gb": 0, "blocked": []}
    segments = [
        segment for segment in db.all(
            "SELECT * FROM recording_segments WHERE session_id=? AND status IN ('complete','analyzed') ORDER BY id",
            (candidate["session_id"],),
        ) if _candidate_overlaps_segment(candidate, segment)
    ]
    return cleanup_ready_segments([int(segment["id"]) for segment in segments])


def cleanup_all_ready_segments(limit: int = 100) -> dict[str, Any]:
    rows = db.all(
        "SELECT id FROM recording_segments WHERE status IN ('complete','analyzed') ORDER BY id LIMIT ?",
        (max(1, min(limit, 1000)),),
    )
    return cleanup_ready_segments([int(row["id"]) for row in rows])


@app.get("/api/review")
def api_review(room_id: str = "", status: str = "pending_review", model: str = "",
               output_date: str = "", export_date: str = "") -> dict[str, Any]:
    selected_room_id = optional_query_int(room_id, "直播间")
    conditions: list[str] = []
    params: list[Any] = []
    if status != "all":
        conditions.append("status=?"); params.append(status)
    if selected_room_id is not None:
        conditions.append("room_id=?"); params.append(selected_room_id)
    if model == "gpt": conditions.append("analysis_version NOT LIKE '%deepseek%'")
    if model == "deepseek": conditions.append("analysis_version LIKE '%deepseek%'")
    date_range = candidate_output_date_range(output_date)
    if date_range:
        conditions.append("created_at>=? AND created_at<?")
        params.extend(date_range)
    exported_range = candidate_output_date_range(export_date)
    if exported_range:
        conditions.append("exported_at>=? AND exported_at<?")
        params.extend(exported_range)
    where = " WHERE " + " AND ".join(conditions) if conditions else ""
    return {"candidates": [candidate_view(row) for row in db.all(
        "SELECT * FROM highlight_candidates" + where + " ORDER BY id DESC", params
    )]}


@app.get("/media")
def serve_media(path: str) -> FileResponse:
    return FileResponse(
        allowed_media(path),
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


class RenderRequest(BaseModel):
    start_time: float
    end_time: float
    caption_text: str = ""
    # Persisted ranges may also carry clause_ids for traceability. Validation
    # below intentionally consumes only start/end.
    source_ranges: list[dict[str, Any]] | None = None


@app.post("/api/candidates/{candidate_id}/render")
def render_candidate(candidate_id: int, payload: RenderRequest) -> dict[str, Any]:
    ranges, duration = validate_ranges(payload.source_ranges or [{"start": payload.start_time, "end": payload.end_time}])
    if not settings.clip_min_seconds <= duration <= settings.clip_max_seconds:
        raise HTTPException(422, f"片段必须为 {settings.clip_min_seconds:.0f}–{settings.clip_max_seconds:.0f} 秒")
    candidate = get_candidate(candidate_id)
    captions = evenly_timed_captions(to_simplified(payload.caption_text), duration)
    version = int(candidate["version"]) + 1
    db.execute(
        """UPDATE highlight_candidates SET start_time=?,end_time=?,captions_json=?,source_ranges_json=?,
           version=?,status='rendering',render_phase='preparing',render_started_at=?,
           render_worker='manual',render_encoder=?,updated_at=? WHERE id=?""",
        (
            float(ranges[0]["start"]), float(ranges[-1]["end"]), json.dumps(captions, ensure_ascii=False),
            json.dumps(ranges, ensure_ascii=False),
            version, utc_now(), media.active_encoder, utc_now(), candidate_id,
        ),
    )
    candidate = get_candidate(candidate_id)
    destination = settings.output_dir / "previews" / f"{safe_id(candidate['source_id'])}_{candidate_id}_v{version}.mp4"
    try:
        def render_progress(phase: str, encoder: str) -> None:
            db.execute(
                """UPDATE highlight_candidates SET render_phase=?,render_encoder=?,updated_at=?
                   WHERE id=? AND status='rendering'""",
                (phase, encoder, utc_now(), candidate_id),
            )

        media.render_candidate(
            db, candidate, captions, destination, progress=render_progress, worker="manual"
        )
    except MediaError as exc:
        db.execute(
            """UPDATE highlight_candidates SET status='render_error',render_phase='failed',
               updated_at=? WHERE id=?""", (utc_now(), candidate_id)
        )
        raise HTTPException(500, str(exc)) from exc
    clean_reason = str(candidate.get("reason") or "").split("；渲染失败：", 1)[0]
    db.execute(
        """UPDATE highlight_candidates
           SET preview_path=?,status='pending_review',reason=?,render_phase='complete',updated_at=? WHERE id=?""",
        (str(destination), clean_reason, utc_now(), candidate_id),
    )
    return {"ok": True, "candidate": get_candidate(candidate_id)}


class ReviewRequest(BaseModel):
    action: Literal["accept", "reject", "defer"]
    reason: str = ""
    start_time: float | None = None
    end_time: float | None = None
    caption_text: str | None = None
    source_ranges: list[dict[str, Any]] | None = None


class BatchReviewRequest(BaseModel):
    candidate_ids: list[int] = Field(min_length=1, max_length=100)
    action: Literal["accept", "reject", "defer"] = "accept"
    reason: str = ""


class BatchExportRequest(BaseModel):
    candidate_ids: list[int] = Field(min_length=1, max_length=100)


@app.post("/api/candidates/{candidate_id}/review")
def review_candidate(candidate_id: int, payload: ReviewRequest) -> dict[str, Any]:
    candidate = get_candidate(candidate_id)
    range_edited = payload.source_ranges is not None or payload.start_time is not None or payload.end_time is not None
    if range_edited:
        ranges, duration = validate_ranges(payload.source_ranges or [{
            "start": float(payload.start_time if payload.start_time is not None else candidate["start_time"]),
            "end": float(payload.end_time if payload.end_time is not None else candidate["end_time"]),
        }])
        start, end = float(ranges[0]["start"]), float(ranges[-1]["end"])
        if not settings.clip_min_seconds <= duration <= settings.clip_max_seconds:
            raise HTTPException(422, "审核片段必须保持在15–20秒")
    else:
        # Accept/reject/defer are status decisions. They must remain available
        # for older candidates even when their stored model ranges no longer
        # satisfy today's manual editing rules.
        ranges = candidate["source_ranges"] or [{
            "start": float(candidate["start_time"]), "end": float(candidate["end_time"]),
        }]
        start, end = float(candidate["start_time"]), float(candidate["end_time"])
        duration = sum(max(0.0, float(item["end"]) - float(item["start"])) for item in ranges)
    captions = candidate["captions"]
    if payload.caption_text is not None:
        captions = evenly_timed_captions(to_simplified(payload.caption_text), duration)
    desired_status = {"accept": "accepted", "reject": "rejected", "defer": "deferred"}[payload.action]
    idempotent_statuses = {"accept": {"accepted", "exported"}, "reject": {"rejected"}, "defer": {"deferred"}}
    if (candidate["status"] in idempotent_statuses[payload.action]
            and start == candidate["start_time"] and end == candidate["end_time"]
            and captions == candidate["captions"]):
        auto_cleanup = cleanup_ready_segments_for_candidate(candidate_id) if payload.action == "reject" else None
        return {"ok": True, "candidate": candidate, "idempotent": True, "auto_cleanup": auto_cleanup}
    status = candidate["status"] if payload.action == "accept" and candidate["status"] == "exported" else desired_status
    version = int(candidate["version"]) + int(start != candidate["start_time"] or end != candidate["end_time"] or captions != candidate["captions"])
    db.execute(
        """UPDATE highlight_candidates SET status=?,start_time=?,end_time=?,captions_json=?,source_ranges_json=?,
           catalog_item_id=?,version=?,updated_at=? WHERE id=?""",
        (status, start, end, json.dumps(captions, ensure_ascii=False), json.dumps(ranges, ensure_ascii=False), None, version, utc_now(), candidate_id),
    )
    db.execute(
        """INSERT INTO review_decisions
           (candidate_id,action,reason,start_time,end_time,captions_json,catalog_item_id,candidate_version,created_at)
           VALUES(?,?,?,?,?,?,?,?,?)""",
        (candidate_id, payload.action, payload.reason, start, end, json.dumps(captions, ensure_ascii=False), None, version, utc_now()),
    )
    updated = get_candidate(candidate_id)
    db.execute(
        "UPDATE publish_jobs SET status='cancelled',updated_at=? WHERE candidate_id=? AND status NOT IN ('exported','published')",
        (utc_now(), candidate_id),
    )
    auto_cleanup = cleanup_ready_segments_for_candidate(candidate_id) if payload.action == "reject" else None
    return {"ok": True, "candidate": updated, "auto_cleanup": auto_cleanup}


@app.post("/api/review/batch")
def batch_review_candidates(payload: BatchReviewRequest) -> dict[str, Any]:
    candidate_ids = list(dict.fromkeys(payload.candidate_ids))
    candidates = [get_candidate(candidate_id) for candidate_id in candidate_ids]
    for candidate in candidates:
        allowed_statuses = {"pending_review", "accepted", "exported"} if payload.action == "accept" else {"pending_review", "deferred"}
        if candidate["status"] not in allowed_statuses:
            raise HTTPException(409, f"候选 #{candidate['id']} 当前状态不可批量审核")
    results = []
    for candidate in candidates:
        result = review_candidate(candidate["id"], ReviewRequest(
            action=payload.action,
            reason=payload.reason,
        ))
        results.append(result["candidate"])
    return {"ok": True, "count": len(results), "candidate_ids": candidate_ids}


@app.post("/api/candidates/{candidate_id}/export")
def export_candidate(candidate_id: int) -> dict[str, Any]:
    candidate = get_candidate(candidate_id)
    if candidate["status"] not in {"accepted", "exported"}:
        raise HTTPException(409, "只有已接受的片段可以导出")
    if candidate["status"] == "exported" and candidate.get("output_path") and Path(candidate["output_path"]).exists():
        destination = Path(candidate["output_path"])
        metadata_path = str(destination.with_suffix(".json"))
        auto_cleanup = cleanup_ready_segments_for_candidate(candidate_id)
        return {"ok": True, "output_path": candidate["output_path"], "metadata_path": metadata_path,
                "exported_at": candidate.get("exported_at") or "", "idempotent": True,
                "auto_cleanup": auto_cleanup}
    stem = f"{safe_id(candidate['source_id'])}_{safe_id(candidate['session_id'])}_{candidate_id}"
    destination = settings.output_dir / "approved" / f"{stem}.mp4"
    preview_path = Path(candidate.get("preview_path") or "")
    reused_preview = preview_path.is_file()
    if reused_preview:
        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.unlink(missing_ok=True)
        try:
            # A hard link is effectively instant and consumes no duplicate
            # disk space. Both files live under the same portable build.
            os.link(preview_path, destination)
        except OSError:
            shutil.copy2(preview_path, destination)
    else:
        try:
            media.render_candidate(db, candidate, candidate["captions"], destination)
        except MediaError as exc:
            raise HTTPException(500, str(exc)) from exc
    exported_at = utc_now()
    metadata = {
        "candidate_id": candidate_id,
        "source_id": candidate["source_id"],
        "session_id": candidate["session_id"],
        "source_start": candidate["start_time"],
        "source_end": candidate["end_time"],
        "source_ranges": candidate["source_ranges"],
        "duration": candidate["duration"],
        "scores": {
            "sales": candidate["sales_score"], "coherence": candidate["coherence_score"],
            "product": candidate["product_score"], "confidence": candidate["confidence"],
        },
        "reason": candidate["reason"],
        "risks": candidate["risks"],
        "kept_clauses": candidate["kept_clauses"],
        "removed_clauses": candidate["removed_clauses"],
        "compliance_hits": candidate["compliance_hits"],
        "versions": {"analysis": candidate["analysis_version"], "prompt": candidate["prompt_version"], "rules": candidate["rule_version"]},
        "captions": candidate["captions"],
        "export_source": "existing_preview" if reused_preview else "rendered_from_source",
        "exported_at": exported_at,
    }
    metadata_path = destination.with_suffix(".json")
    metadata_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    db.execute(
        """UPDATE highlight_candidates SET output_path=?,status='exported',exported_at=?,
           media_cleaned_at='',media_released_bytes=0,updated_at=? WHERE id=?""",
        (str(destination), exported_at, exported_at, candidate_id),
    )
    auto_cleanup = cleanup_ready_segments_for_candidate(candidate_id)
    return {"ok": True, "output_path": str(destination), "metadata_path": str(metadata_path),
            "exported_at": exported_at, "idempotent": False, "auto_cleanup": auto_cleanup}


@app.post("/api/candidates/batch-export")
def batch_export_candidates(payload: BatchExportRequest) -> dict[str, Any]:
    candidate_ids = list(dict.fromkeys(payload.candidate_ids))
    candidates = [get_candidate(candidate_id) for candidate_id in candidate_ids]
    for candidate in candidates:
        if candidate["status"] not in {"accepted", "exported"}:
            raise HTTPException(409, f"候选 #{candidate['id']} 尚未审核接受，不能导出")

    exported_files: list[tuple[dict[str, Any], Path]] = []
    for candidate in candidates:
        result = export_candidate(int(candidate["id"]))
        output_path = Path(result["output_path"])
        if not output_path.is_file():
            raise HTTPException(500, f"候选 #{candidate['id']} 的成片生成失败")
        exported_files.append((get_candidate(int(candidate["id"])), output_path))

    batch_dir = settings.cache_dir / "batch_exports"
    batch_dir.mkdir(parents=True, exist_ok=True)
    stale_before = datetime.now(timezone.utc).timestamp() - 3600
    for stale_archive in batch_dir.glob("*.zip"):
        try:
            if stale_archive.stat().st_mtime < stale_before:
                stale_archive.unlink()
        except OSError:
            pass
    token = uuid4().hex
    archive_path = batch_dir / f"{token}.zip"
    manifest = {
        "exported_at": utc_now(),
        "count": len(exported_files),
        "candidate_ids": candidate_ids,
        "items": [],
    }
    room_prefixes = {candidate_room_file_prefix(candidate) for candidate, _ in exported_files}
    output_days = {candidate_local_date(candidate, True) for candidate, _ in exported_files}
    package_room = next(iter(room_prefixes)) if len(room_prefixes) == 1 else "多个直播间"
    package_day = next(iter(output_days)) if len(output_days) == 1 else "多日期"
    download_name = f"{package_room}_{package_day}_{len(exported_files)}条素材.zip"
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_STORED) as archive:
        for candidate, output_path in exported_files:
            room_prefix = candidate_room_file_prefix(candidate)
            archive_name = candidate_download_filename(candidate)
            if len(room_prefixes) > 1:
                archive_name = f"{room_prefix}/{archive_name}"
            archive.write(output_path, archive_name)
            manifest["items"].append({
                "candidate_id": candidate["id"],
                "room": candidate.get("room_name") or candidate.get("source_id") or "",
                "output_date": candidate.get("created_at") or "",
                "exported_at": candidate.get("exported_at") or "",
                "filename": archive_name,
            })
        archive.writestr("导出清单.json", json.dumps(manifest, ensure_ascii=False, indent=2))
    return {
        "ok": True,
        "count": len(exported_files),
        "filename": download_name,
        "download_url": f"/api/batch-exports/{token}?count={len(exported_files)}&filename={quote(download_name)}",
    }


@app.get("/api/batch-exports/{token}")
def download_batch_export(token: str, count: int = 0, filename: str = "") -> FileResponse:
    if len(token) != 32 or any(char not in "0123456789abcdef" for char in token.lower()):
        raise HTTPException(404, "批量导出文件不存在")
    archive_path = settings.cache_dir / "batch_exports" / f"{token.lower()}.zip"
    if not archive_path.is_file():
        raise HTTPException(404, "批量导出文件不存在或已完成清理")
    stamp = datetime.now(timezone(timedelta(hours=8))).strftime("%Y%m%d_%H%M%S")
    requested_stem = safe_id(Path(filename).stem) if filename else ""
    download_name = f"{requested_stem}.zip" if requested_stem else f"直播成片_{stamp}_{max(1, count)}条.zip"
    return FileResponse(
        archive_path,
        media_type="application/zip",
        filename=download_name,
        background=BackgroundTask(archive_path.unlink, missing_ok=True),
    )


@app.get("/api/candidates/{candidate_id}/download")
def download_candidate(candidate_id: int) -> FileResponse:
    candidate = get_candidate(candidate_id)
    if candidate["status"] != "exported" or not candidate.get("output_path"):
        raise HTTPException(409, "该成片尚未导出，或本机文件已按保留期清理")
    path = allowed_media(candidate["output_path"])
    return FileResponse(path, media_type="video/mp4", filename=candidate_download_filename(candidate))


@app.get("/publish", response_class=HTMLResponse)
def publish_page(request: Request, status: str = "all") -> HTMLResponse:
    return RedirectResponse("/", status_code=302)
    cleanup_expired_candidate_media()
    if status == "all":
        where, params = "", ()
    elif status == "handed_off":
        where, params = " WHERE p.status='exported' AND p.handoff_confirmed_at<>''", ()
    elif status == "exported":
        where, params = " WHERE p.status='exported' AND p.handoff_confirmed_at=''", ()
    else:
        where, params = " WHERE p.status=?", (status,)
    jobs = db.all(
        f"""SELECT p.*,h.preview_path,h.output_path,h.source_id,h.media_cleaned_at,h.media_released_bytes,
            r.sequence AS room_sequence,r.name AS room_name,
            c.name AS catalog_name,
            (SELECT MAX(d.created_at) FROM review_decisions d
             WHERE d.candidate_id=p.candidate_id AND d.action='accept') AS accepted_at
            FROM publish_jobs p
            JOIN highlight_candidates h ON h.id=p.candidate_id
            LEFT JOIN live_rooms r ON r.id=p.room_id LEFT JOIN catalog_items c ON c.id=p.catalog_item_id
            {where} ORDER BY p.id DESC LIMIT 300""", params,
    )
    groups: dict[str, dict[str, Any]] = {}
    for job in jobs:
        key = f"catalog-{job['catalog_item_id']}" if job.get("catalog_item_id") else f"missing-{job['id']}"
        group = groups.setdefault(key, {
            "key": key, "catalog_item_id": job.get("catalog_item_id"),
            "internal_code": job.get("internal_code_snapshot") or "未关联款号",
            "catalog_name": job.get("catalog_name") or "投放信息待补",
            "qianchuan_product_id": job.get("qianchuan_product_id_snapshot") or "",
            "qianchuan_plan_id": job.get("qianchuan_plan_id_snapshot") or "",
            "jobs": [], "rooms": set(), "accepted_at": "", "status": "exported",
            "package_path": "", "handoff_confirmed": True,
            "local_media_count": 0, "local_media_bytes": 0,
        })
        group["jobs"].append(job)
        group["rooms"].add(f"{job.get('room_sequence') or '-'} · {job.get('room_name') or job.get('source_id')}")
        if str(job.get("accepted_at") or "") > group["accepted_at"]:
            group["accepted_at"] = job.get("accepted_at") or ""
        if job.get("package_path"):
            group["package_path"] = job["package_path"]
        if not job.get("handoff_confirmed_at"):
            group["handoff_confirmed"] = False
        if not job.get("media_cleaned_at"):
            local_paths = {str(job.get("preview_path") or ""), str(job.get("output_path") or "")} - {""}
            sizes = 0
            for path_text in local_paths:
                path = Path(path_text)
                if path.is_file():
                    sizes += path.stat().st_size
            if sizes:
                group["local_media_count"] += 1
                group["local_media_bytes"] += sizes
        if job["status"] == "missing_info": group["status"] = "missing_info"
        elif job["status"] == "ready" and group["status"] != "missing_info": group["status"] = "ready"
        elif job["status"] == "cancelled" and group["status"] == "exported": group["status"] = "cancelled"
    for group in groups.values():
        group["rooms"] = sorted(group["rooms"])
        group["candidate_ids"] = [job["candidate_id"] for job in group["jobs"]]
        group["job_ids"] = [job["id"] for job in group["jobs"] if job["status"] != "cancelled"]
        group["video_count"] = len(group["job_ids"])
        group["local_media_mb"] = round(group["local_media_bytes"] / (1024 ** 2), 1)
        reports = handoff_segment_reports(group["job_ids"]) if group["job_ids"] else []
        group["raw_segment_count"] = sum(not report["already_cleaned"] for report in reports)
        group["raw_cleanable"] = bool(reports) and all(
            report["cleanable"] or report["already_cleaned"] for report in reports
        )
        group["raw_estimated_gb"] = round(sum(
            report["estimated_bytes"] for report in reports if not report["already_cleaned"]
        ) / (1024 ** 3), 3)
        if group["status"] == "exported" and group["handoff_confirmed"]:
            group["status"] = "handed_off"
    return templates.TemplateResponse(request, "publish.html", {
        "groups": list(groups.values()), "jobs": jobs, "selected_status": status,
        "catalog": db.all("SELECT * FROM catalog_items WHERE active=1 ORDER BY internal_code LIMIT 500"),
    })


class PackageExportRequest(BaseModel):
    job_ids: list[int] = Field(min_length=1, max_length=100)


@app.post("/api/handoff-packages/export")
def export_handoff_package(payload: PackageExportRequest) -> dict[str, Any]:
    def china_time(value: str) -> str:
        if not value:
            return "未记录"
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=timezone.utc)
            return parsed.astimezone(timezone(timedelta(hours=8))).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            return value

    job_ids = list(dict.fromkeys(payload.job_ids))
    placeholders = ",".join("?" for _ in job_ids)
    jobs = db.all(
        f"""SELECT p.*,h.preview_path,h.output_path,h.source_id,
            r.sequence AS room_sequence,r.name AS room_name,c.name AS catalog_name,
            (SELECT MAX(d.created_at) FROM review_decisions d
             WHERE d.candidate_id=p.candidate_id AND d.action='accept') AS accepted_at
            FROM publish_jobs p
            JOIN highlight_candidates h ON h.id=p.candidate_id
            LEFT JOIN live_rooms r ON r.id=p.room_id
            LEFT JOIN catalog_items c ON c.id=p.catalog_item_id
            WHERE p.id IN ({placeholders}) ORDER BY p.id""", job_ids
    )
    if len(jobs) != len(job_ids):
        raise HTTPException(404, "部分人工交接任务不存在")
    catalog_ids = {job.get("catalog_item_id") for job in jobs}
    if len(catalog_ids) != 1 or None in catalog_ids:
        raise HTTPException(422, "一个压缩包只能包含同一个已关联款号")
    first = jobs[0]
    if not first.get("internal_code_snapshot") or not first.get("qianchuan_product_id_snapshot"):
        raise HTTPException(409, "款号或千川商品ID尚未填写完整")
    if any(job["status"] not in {"ready", "exported"} for job in jobs):
        raise HTTPException(409, "所选任务中包含不可导出的状态")

    package_dir = settings.output_dir / "packages"
    package_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    code = safe_id(first["internal_code_snapshot"])
    destination = package_dir / f"{code}_{len(jobs)}条视频_{stamp}.zip"
    video_items: list[tuple[dict[str, Any], Path, str]] = []
    for job in jobs:
        output_path = Path(job["output_path"]) if job.get("output_path") else None
        preview_path = Path(job["preview_path"]) if job.get("preview_path") else None
        source_path = output_path if output_path and output_path.is_file() else preview_path
        if not source_path or not source_path.is_file():
            raise HTTPException(409, f"候选 #{job['candidate_id']} 的成片不存在，请先重新渲染")
        archive_name = f"{code}_候选{job['candidate_id']}{source_path.suffix.lower() or '.mp4'}"
        video_items.append((job, source_path, archive_name))

    packaged_at = utc_now()
    info_lines = [
        "人工审核视频交接信息",
        "====================",
        f"款号：{first['internal_code_snapshot']}",
        f"商品名称：{first.get('catalog_name') or '未填写'}",
        f"千川商品ID：{first['qianchuan_product_id_snapshot']}",
        f"千川计划ID：{first.get('qianchuan_plan_id_snapshot') or '未填写'}",
        f"视频数量：{len(video_items)}",
        f"打包时间（北京时间）：{china_time(packaged_at)}",
        "",
        "视频明细",
        "--------",
    ]
    for index, (job, _source_path, archive_name) in enumerate(video_items, start=1):
        room = " · ".join(filter(None, [job.get("room_sequence"), job.get("room_name")]))
        info_lines.extend([
            f"{index}. 文件：{archive_name}",
            f"   候选编号：#{job['candidate_id']}",
            f"   来源直播间：{room or job.get('source_id') or '未记录'}",
            f"   人工审核时间（北京时间）：{china_time(job.get('accepted_at') or '')}",
        ])
    info_text = "\n".join(info_lines) + "\n"

    # MP4/H.264 本身已经压缩过，再压缩几乎不省空间，却会显著拖慢交接。
    with zipfile.ZipFile(destination, "w", compression=zipfile.ZIP_STORED) as archive:
        for _job, source_path, archive_name in video_items:
            archive.write(source_path, arcname=archive_name)
        archive.writestr("人工交接信息.txt", info_text)

    for job, source_path, _archive_name in video_items:
        db.execute(
            "UPDATE highlight_candidates SET output_path=?,status='exported',updated_at=? WHERE id=?",
            (str(source_path), packaged_at, job["candidate_id"]),
        )
        db.execute(
            """UPDATE publish_jobs SET status='exported',error='',package_path=?,
               handoff_confirmed_at='',updated_at=? WHERE id=?""",
            (str(destination), packaged_at, job["id"]),
        )
    return {"ok": True, "package_path": str(destination), "filename": destination.name,
            "video_count": len(jobs)}


class HandoffConfirmRequest(BaseModel):
    job_ids: list[int] = Field(min_length=1, max_length=100)


@app.post("/api/handoff-packages/confirm")
def confirm_handoff_package(payload: HandoffConfirmRequest) -> dict[str, Any]:
    job_ids = list(dict.fromkeys(payload.job_ids))
    placeholders = ",".join("?" for _ in job_ids)
    jobs = db.all(f"SELECT * FROM publish_jobs WHERE id IN ({placeholders})", job_ids)
    if len(jobs) != len(job_ids):
        raise HTTPException(404, "部分人工交接任务不存在")
    package_paths = {job.get("package_path") for job in jobs}
    if len(package_paths) != 1 or not next(iter(package_paths), ""):
        raise HTTPException(409, "请先把这一批素材打包")
    package_path = Path(next(iter(package_paths)))
    if not package_path.is_file():
        raise HTTPException(409, "交接压缩包不存在，请重新打包")
    try:
        with zipfile.ZipFile(package_path, "r") as archive:
            if archive.testzip() is not None:
                raise HTTPException(409, "压缩包校验失败，请重新打包")
    except zipfile.BadZipFile as exc:
        raise HTTPException(409, "压缩包无法打开，请重新打包") from exc
    confirmed_at = utc_now()
    db.execute_many(
        "UPDATE publish_jobs SET handoff_confirmed_at=?,updated_at=? WHERE id=?",
        [(confirmed_at, confirmed_at, job_id) for job_id in job_ids],
    )
    return {"ok": True, "count": len(job_ids), "confirmed_at": confirmed_at}


@app.post("/api/handoff-packages/cleanup-media")
def cleanup_handoff_media(payload: HandoffConfirmRequest) -> dict[str, Any]:
    job_ids = list(dict.fromkeys(payload.job_ids))
    placeholders = ",".join("?" for _ in job_ids)
    jobs = db.all(f"SELECT * FROM publish_jobs WHERE id IN ({placeholders})", job_ids)
    if len(jobs) != len(job_ids):
        raise HTTPException(404, "部分人工交接任务不存在")
    if any(not job.get("handoff_confirmed_at") for job in jobs):
        raise HTTPException(409, "只有确认已经取走的素材才能删除本地候选成片")
    released = 0
    for job in jobs:
        result = cleanup_candidate_media(int(job["candidate_id"]))
        released += int(result["released_bytes"])
    return {"ok": True, "count": len(jobs), "released_bytes": released,
            "released_mb": round(released / (1024 ** 2), 1)}


@app.post("/api/handoff-packages/cleanup-segments")
def cleanup_handoff_segments(payload: HandoffConfirmRequest) -> dict[str, Any]:
    job_ids = list(dict.fromkeys(payload.job_ids))
    placeholders = ",".join("?" for _ in job_ids)
    jobs = db.all(f"SELECT * FROM publish_jobs WHERE id IN ({placeholders})", job_ids)
    if len(jobs) != len(job_ids):
        raise HTTPException(404, "部分人工交接任务不存在")
    if any(not job.get("handoff_confirmed_at") for job in jobs):
        raise HTTPException(409, "请先确认这批素材已经下载并取走")
    reports = handoff_segment_reports(job_ids)
    blockers = [
        f"分片 #{report['segment']['id']}：" + "；".join(report["blockers"])
        for report in reports if not report["cleanable"] and not report["already_cleaned"]
    ]
    if blockers:
        raise HTTPException(409, {"message": "仍有候选没有明确去向，暂不能清理原始录像", "blockers": blockers})
    released = count = 0
    for report in reports:
        if report["already_cleaned"]:
            continue
        result = cleanup_segment(int(report["segment"]["id"]), SegmentCleanupRequest(confirm=True))
        count += 1
        released += int(result["released_bytes"])
    return {"ok": True, "count": count, "released_bytes": released,
            "released_gb": round(released / (1024 ** 3), 3)}


@app.get("/api/publish-jobs")
def publish_jobs(status: str = "all") -> dict[str, Any]:
    if status == "all":
        rows = db.all("SELECT * FROM publish_jobs ORDER BY id DESC")
    else:
        rows = db.all("SELECT * FROM publish_jobs WHERE status=? ORDER BY id DESC", (status,))
    return {"jobs": rows}


class PublishJobCatalogRequest(BaseModel):
    catalog_item_id: int | None = None
    internal_code: str = ""
    name: str = ""
    qianchuan_product_id: str = ""
    qianchuan_plan_id: str = ""


@app.post("/api/publish-jobs/{job_id}/catalog")
def complete_publish_job_catalog(job_id: int, payload: PublishJobCatalogRequest) -> dict[str, Any]:
    job = db.one("SELECT * FROM publish_jobs WHERE id=?", (job_id,))
    if not job:
        raise HTTPException(404, "人工交接任务不存在")
    if payload.catalog_item_id is not None:
        item = db.one("SELECT * FROM catalog_items WHERE id=? AND active=1", (payload.catalog_item_id,))
        if not item:
            raise HTTPException(422, "所选款号不存在或已停用")
    else:
        code = payload.internal_code.strip()
        product_id = payload.qianchuan_product_id.strip()
        plan_id = payload.qianchuan_plan_id.strip()
        if not code or not product_id or not plan_id:
            raise HTTPException(422, "款号、千川商品ID和千川计划ID必须填写完整")
        result = upsert_catalog(CatalogRequest(
            internal_code=code,
            name=payload.name.strip() or code,
            qianchuan_product_id=product_id,
            qianchuan_plan_id=plan_id,
        ))
        item = result["item"]
    if not item.get("qianchuan_product_id") or not item.get("qianchuan_plan_id"):
        raise HTTPException(422, "这个款号的千川商品ID或计划ID尚未填写完整")
    now = utc_now()
    db.execute(
        "UPDATE highlight_candidates SET catalog_item_id=?,updated_at=? WHERE id=?",
        (item["id"], now, job["candidate_id"]),
    )
    updated = ensure_publish_job(get_candidate(int(job["candidate_id"])))
    db.execute(
        """INSERT INTO review_decisions
           (candidate_id,action,reason,start_time,end_time,captions_json,catalog_item_id,candidate_version,created_at)
           SELECT id,'catalog_update','在人工交接页补充投放信息',start_time,end_time,captions_json,?,version,?
           FROM highlight_candidates WHERE id=?""",
        (item["id"], now, job["candidate_id"]),
    )
    return {"ok": True, "job": updated, "item": item}


@app.post("/api/publish-jobs/{job_id}/export")
def export_publish_job(job_id: int) -> dict[str, Any]:
    job = db.one("SELECT * FROM publish_jobs WHERE id=?", (job_id,))
    if not job:
        raise HTTPException(404, "人工交接单不存在")
    if job["status"] == "missing_info":
        raise HTTPException(409, "款号、千川商品ID或计划ID不完整")
    return export_candidate(int(job["candidate_id"]))


@app.post("/api/publish-jobs/{job_id}/cancel")
def cancel_publish_job(job_id: int) -> dict[str, Any]:
    job = db.one("SELECT * FROM publish_jobs WHERE id=?", (job_id,))
    if not job:
        raise HTTPException(404, "人工交接单不存在")
    if job["status"] == "published":
        raise HTTPException(409, "已完成交接记录不能取消")
    db.execute("UPDATE publish_jobs SET status='cancelled',updated_at=? WHERE id=?", (utc_now(), job_id))
    return {"ok": True}


class CatalogRequest(BaseModel):
    internal_code: str = Field(min_length=1, max_length=100)
    name: str = Field(min_length=1, max_length=200)
    aliases: str = ""
    qianchuan_product_id: str = ""
    qianchuan_plan_id: str = ""
    reference_images: str = ""
    notes: str = ""
    active: bool = True


@app.post("/api/catalog")
def upsert_catalog(payload: CatalogRequest) -> dict[str, Any]:
    now = utc_now()
    existing = db.one("SELECT id FROM catalog_items WHERE internal_code=?", (payload.internal_code.strip(),))
    values = (
        payload.name.strip(), payload.aliases.strip(), payload.qianchuan_product_id.strip(),
        payload.qianchuan_plan_id.strip(), payload.reference_images.strip(), payload.notes.strip(),
        int(payload.active), now,
    )
    if existing:
        db.execute(
            """UPDATE catalog_items SET name=?,aliases=?,qianchuan_product_id=?,qianchuan_plan_id=?,
               reference_images=?,notes=?,active=?,updated_at=? WHERE id=?""",
            (*values, existing["id"]),
        )
        item_id = existing["id"]
    else:
        item_id = db.execute(
            """INSERT INTO catalog_items
               (internal_code,name,aliases,qianchuan_product_id,qianchuan_plan_id,reference_images,notes,active,created_at,updated_at)
               VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (payload.internal_code.strip(), *values[:-1], now, now),
        )
    return {"ok": True, "item": db.one("SELECT * FROM catalog_items WHERE id=?", (item_id,))}


@app.delete("/api/catalog/{item_id}")
def delete_catalog(item_id: int, purge_history: bool = False) -> dict[str, Any]:
    item = db.one("SELECT * FROM catalog_items WHERE id=?", (item_id,))
    if not item:
        raise HTTPException(404, "款号不存在")
    usage = {
        "segments": int(db.one("SELECT COUNT(*) count FROM recording_segments WHERE catalog_item_id=?", (item_id,))["count"]),
        "candidates": int(db.one("SELECT COUNT(*) count FROM highlight_candidates WHERE catalog_item_id=?", (item_id,))["count"]),
        "handoffs": int(db.one("SELECT COUNT(*) count FROM publish_jobs WHERE catalog_item_id=?", (item_id,))["count"]),
        "reviews": int(db.one("SELECT COUNT(*) count FROM review_decisions WHERE catalog_item_id=?", (item_id,))["count"]),
    }
    now = utc_now()
    has_history = any(usage.values())
    if purge_history and has_history:
        candidates = db.all(
            """SELECT DISTINCT h.* FROM highlight_candidates h
               LEFT JOIN publish_jobs p ON p.candidate_id=h.id
               WHERE h.catalog_item_id=? OR p.catalog_item_id=?""", (item_id, item_id),
        )
        candidate_ids = [int(candidate["id"]) for candidate in candidates]
        if candidate_ids:
            candidate_placeholders = ",".join("?" for _ in candidate_ids)
            package_rows = db.all(
                f"""SELECT package_path FROM publish_jobs
                    WHERE package_path<>'' AND (catalog_item_id=? OR candidate_id IN ({candidate_placeholders}))""",
                (item_id, *candidate_ids),
            )
        else:
            package_rows = db.all(
                "SELECT package_path FROM publish_jobs WHERE catalog_item_id=? AND package_path<>''", (item_id,)
            )
        package_paths = {row["package_path"] for row in package_rows}
        released = 0
        data_roots = [settings.output_dir.resolve(), settings.cache_dir.resolve(), settings.keyframe_dir.resolve()]
        for candidate in candidates:
            seen_paths: set[str] = set()
            for field in ("preview_path", "output_path"):
                path_text = str(candidate.get(field) or "")
                if path_text and path_text not in seen_paths:
                    released += _unlink_cleanup_file(Path(path_text), data_roots)
                    seen_paths.add(path_text)
            for path in settings.keyframe_dir.glob(f"candidate_{candidate['id']}_*"):
                released += _unlink_cleanup_file(path, data_roots)
            for pattern in (f"candidate_{candidate['id']}.*", f"candidate_{candidate['id']}_*"):
                for path in settings.cache_dir.glob(pattern):
                    released += _unlink_cleanup_file(path, data_roots)
        db.execute("UPDATE live_rooms SET default_catalog_item_id=NULL,updated_at=? WHERE default_catalog_item_id=?", (now, item_id))
        db.execute("UPDATE recording_segments SET catalog_item_id=NULL,updated_at=? WHERE catalog_item_id=?", (now, item_id))
        db.execute("DELETE FROM publish_jobs WHERE catalog_item_id=?", (item_id,))
        if candidate_ids:
            placeholders = ",".join("?" for _ in candidate_ids)
            db.execute(f"DELETE FROM highlight_candidates WHERE id IN ({placeholders})", candidate_ids)
        db.execute("DELETE FROM review_decisions WHERE catalog_item_id=?", (item_id,))
        db.execute("DELETE FROM catalog_items WHERE id=?", (item_id,))
        for package_path in package_paths:
            if not db.one("SELECT id FROM publish_jobs WHERE package_path=? LIMIT 1", (package_path,)):
                released += _unlink_cleanup_file(Path(package_path), [settings.output_dir.resolve()])
        return {"ok": True, "mode": "purged", "usage": usage,
                "released_bytes": released, "released_mb": round(released / (1024 ** 2), 1),
                "message": "款号及关联测试记录已彻底删除"}

    db.execute("UPDATE live_rooms SET default_catalog_item_id=NULL,updated_at=? WHERE default_catalog_item_id=?", (now, item_id))
    if not has_history:
        db.execute("DELETE FROM catalog_items WHERE id=?", (item_id,))
        mode = "deleted"
    else:
        db.execute("UPDATE catalog_items SET active=0,updated_at=? WHERE id=?", (now, item_id))
        mode = "deactivated"
    return {"ok": True, "mode": mode, "usage": usage,
            "message": "款号已删除" if mode == "deleted" else "款号已从ID库移除，历史交接记录已保留"}


CATALOG_HEADERS = ["internal_code", "name", "aliases", "qianchuan_product_id", "qianchuan_plan_id", "reference_images", "notes", "active"]


@app.post("/api/catalog/import")
async def import_catalog(file: UploadFile = File(...)) -> dict[str, Any]:
    content = await file.read()
    rows: list[dict[str, Any]] = []
    if (file.filename or "").lower().endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if values:
            headers = [str(value or "").strip() for value in values[0]]
            rows = [dict(zip(headers, row)) for row in values[1:]]
    else:
        decoded = content.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(decoded)))
    imported = 0
    for row in rows:
        code = str(row.get("internal_code") or "").strip()
        name = str(row.get("name") or "").strip()
        if not code or not name:
            continue
        active_value = str(row.get("active", "1")).strip().lower()
        upsert_catalog(CatalogRequest(
            internal_code=code, name=name, aliases=str(row.get("aliases") or ""),
            qianchuan_product_id=str(row.get("qianchuan_product_id") or ""),
            qianchuan_plan_id=str(row.get("qianchuan_plan_id") or ""),
            reference_images=str(row.get("reference_images") or ""), notes=str(row.get("notes") or ""),
            active=active_value not in {"0", "false", "否", "no"},
        ))
        imported += 1
    return {"ok": True, "imported": imported}


@app.get("/api/catalog/export")
def export_catalog() -> Response:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品ID库"
    sheet.append(CATALOG_HEADERS)
    for row in db.all("SELECT * FROM catalog_items ORDER BY internal_code"):
        sheet.append([row.get(header, "") for header in CATALOG_HEADERS])
    buffer = io.BytesIO()
    workbook.save(buffer)
    headers = {"Content-Disposition": "attachment; filename=catalog.xlsx"}
    return Response(buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@app.post("/api/segments/{segment_id}/retry")
def retry_segment(segment_id: int) -> dict[str, Any]:
    try:
        pipeline.retry_segment(segment_id)
    except KeyError as exc:
        raise HTTPException(404, "分片不存在") from exc
    return {"ok": True}


def run() -> None:
    import uvicorn

    uvicorn.run("app.main:app", host=settings.host, port=settings.port, reload=False)


if __name__ == "__main__":
    run()
